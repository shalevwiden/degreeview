import requests
import bs4
from bs4 import BeautifulSoup

import sys
import os
import subprocess
import random

import csv
import openpyxl
import json




from openpyxl import load_workbook
from openpyxl import Workbook
# use this to define the function define_start_column
from openpyxl.utils import column_index_from_string

# this assigns cells colors
from openpyxl.styles import PatternFill

import time
import os

from openpyxl.styles import Font
from openpyxl.styles import Border, Side, Alignment

# for working with pdfs:
import cairosvg
from cairosvg import svg2pdf

# png to pdf conversion
from PIL import Image


from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.colors import Color


# for merging the pdfs: and reading them to get data. 
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from pathlib import Path  


# for database stuff
import sqlite3


# Define the absolute path to the folder containing the file you want to import
# Replace this with your actual absolute path
pdftemplatesfolder = Path('/Users/shalevwiden/Downloads/Coding_Files/Python/BeautifulSoup_Library/PDF_Templates')

# Add the folder to the front of sys.path so Python will search here first
sys.path.insert(0, str(pdftemplatesfolder))

# Now you can import the function as if it were local
from stareoverlay import create_stare_png
exceltemplatespath = Path("/Users/shalevwiden/Downloads/Projects/dvwebsitecreation/sourcefiles/pythonfunctions")
sys.path.append(str(exceltemplatespath))

from excel_templates import make_degreeplan_excel_files 


# good to check everythings working with the venv:
if __name__=='__main__':
    print(f'the version of beautiful soup is\n {(bs4.__version__)}')
    print(f'the version of requests is\n {(requests.__version__)}')
    print(f'the version of openpyxl is\n {(openpyxl.__version__)}')

    print(f'\nthe python version being used is:{sys.executable}\n')


# this script is gonna be.. huge. Yeah this is the behemoth one
# the asset is a LIST of dictionaries

with open('theassetcontainment.json') as assetjson:
    theasset=json.load(assetjson)

        
from scrapedata_orclasses import getallcourses_splitbysemester


class makeSemesterFiles:
    '''
    Each object of this class should take an item (list item) from the asset.
    The asset of 0-14 for each school. 
    For testing we will use the asset of 0.
   
    '''
    # school data is like theasset[0], which is what each class object takes
    def __init__(self,schooldata):

        # you can print every attribute in the class which is the superpower tho. 

        # use self to make things an attribute
        self.universityname='The University of Texas at Austin'

        self.schooldata=schooldata
        
        self.schoolnamekey=list(schooldata)[0]
        self.schoolname=schooldata[self.schoolnamekey]

        # this one is fixed 
        self.degreeviewfolderpath='/Users/shalevwiden/Downloads/Projects/originaldegreeview'
        # this should work. If not I need to find a mystery
        self.schoolfolderpath=os.path.join(self.degreeviewfolderpath,self.schoolname)
        
        self.schoolcolordict={'School of Architecture':'#cdbee9', 'Red McCombs School of Business':'#f9e4af', 
 'School of Civic Leadership':'#096c6c', 'Moody College of Communication':'#7e0107',
   'College of Education':'#86dff5', 'Cockrell School of Engineering':'#e35501', 
   'College of Fine Arts':"#633224", 'John A. and Katherine G. Jackson School of Geosciences':'#a3d29c', 
   'School of Information':'#2bc5e0', 
 'College of Liberal Arts':'#f3f3f3', 'College of Natural Sciences':'#f6c44b', 'School of Nursing':'#e3971f',
  'College of Pharmacy':"#BAB86C", 'Lyndon B. Johnson School of Public Affairs':'#096c6c', 'Steve Hicks School of Social Work':'#f0b420'}
        
        self.schoolcolor=self.schoolcolordict[self.schoolname]

        self.logopath='/Users/shalevwiden/Downloads/Projects/dvwebsitecreation/logo_design/logo6.png'


        # this might have to change later lol
        self.websiteurl_img_path='/Users/shalevwiden/Downloads/Projects/dvwebsitecreation/logo_design/websiteurl.png'

        self.make_degreeplan_excel_files=make_degreeplan_excel_files 

        self.configsfolder='/Users/shalevwiden/Downloads/Projects/dvwebsitecreation/sourcefiles/Excelfile_configs'

    def __str__(self):
        returnstring=f'This is an object to create files for {self.schoolname}'
        return returnstring
   
    def finishconfigpath(self,endofpath):
        '''
        This is for configs and is not related to saving the file at all
        '''
        return os.path.join(self.configsfolder,endofpath)
    
    def upload_to_database(self):
        '''
        
        Uses SQLite to upload to databases.
        We will then to quieries on said databases.

        
        '''
        for i in range(1,len(self.schooldata)):
            
            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Starting process for {degreename} ')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            # now use this to write to csv files.
            # I can actually finish this fast
            
            degreefolderpath=os.path.join(self.schoolfolderpath,degreename)
            # can change this quite easily


            # its a little different from the database name, use underscore instead of hyphen
            degreenamecleaned=degreename.replace(' ','').lower().split('(')
            degreenamecleaned=degreenamecleaned[0]+"-"+degreenamecleaned[-1]
            degreenamecleaned=degreenamecleaned.replace(')','')

            

            databasepath=os.path.join(degreefolderpath,f'{degreenamecleaned}-database.db')
            # hyphens and commas not allowed in tablename
            tabledegreename=degreenamecleaned.replace('-','_').replace(',','_').replace('&','and').replace("'","")

            tablename=f'{tabledegreename}_table'
            def maketable():
                '''This creates the table for course data in the db'''
                with sqlite3.connect(databasepath) as conn:
                    cursor=conn.cursor()

                # table name lowercase
                    createtablecommand=f'''CREATE TABLE IF NOT EXISTS "{tablename}" (Coursecode TEXT , Coursename TEXT, Hours INT, Category TEXT, UpperDivStatus TEXT );'''
                    cursor.execute(createtablecommand)
                    conn.commit()
            maketable()
            totalhours=0
            numberofsemesters=len(semesterdictionary)
            

                # .write stuff now mf
                # split up sems 1-4 and 5-8 lol
                
            with sqlite3.connect(databasepath) as conn:
                cursor=conn.cursor()

                    
                for semesternum in range(numberofsemesters):
                    semester=list(semesterdictionary)[semesternum]
                    # semester courses is a dictionary of its own as well
                    semestercourses=semesterdictionary[semester]
                        
                    for coursenameindex in range(len(semestercourses)):

                        coursename=list(semestercourses)[coursenameindex]
                            # if its NOT a list of lists:
                        if len(semestercourses[coursename])==4 and not isinstance(semestercourses[coursename][0],list):
                            coursecode, coursehours, upperdivstatus, coursecategory=semestercourses[coursename]
                            cursor.execute(f'INSERT INTO {tablename} (Coursecode, Coursename, Hours, Category, UpperDivStatus) values(?,?,?,?,?);',
                                   [coursecode,coursename,coursehours,coursecategory,upperdivstatus])
                            

                            if coursehours!='':
                                totalhours+=int(coursehours)
                            
                        else:
                            listofcourses=semestercourses[coursename]
                            for i in range(len(listofcourses)):
                                coursecode, coursehours, upperdivstatus, coursecategory=listofcourses[i]
                                cursor.execute(f'INSERT INTO {tablename} (Coursecode, Coursename, Hours, Category, UpperDivStatus) values(?,?,?,?,?);',
                                   [coursecode,coursename,coursehours,coursecategory,upperdivstatus])

                                if coursehours!='':
                                    totalhours+=int(coursehours)


                    # empty line between semesters
                    cursor.execute(f'INSERT INTO {tablename} (Coursecode, Coursename, Hours, Category, UpperDivStatus) values(?,?,?,?,?);',
                                   [None,None,None,None,None])
                                   
                 # commit to the database 
                conn.commit()
                print(f'Created {degreenamecleaned}-database.db\n')
                





            
            
    def makecsvfiles(self):
        '''
        This uses the result of the scrape semester data function to make a csv file out of it.
        '''

        # for each degree in school data
        for i in range(1,len(self.schooldata)):
            
            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Starting process for {degreename} ')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            # now use this to write to csv files.
            # I can actually finish this fast
            
            degreefolderpath=os.path.join(self.schoolfolderpath,degreename)
            # can change this quite easily


            degreenamecleaned=degreename.replace(' ','').lower().split('(')
            degreenamecleaned=degreenamecleaned[0]+"-"+degreenamecleaned[-1]
            degreenamecleaned=degreenamecleaned.replace(')','')

            semestercsvfilename=os.path.join(degreefolderpath,f'{degreenamecleaned}-semestercsvfilefull.csv')

            totalhours=0
            numberofsemesters=len(semesterdictionary)
            csvobjectdict={}                        
            print(f'\ncsv object has been rest to {len(csvobjectdict)}\n\n\n')
            
            rowcount=0
            with open(semestercsvfilename,'w',newline='') as semestercsvfile:

                # .write stuff now mf
                # split up sems 1-4 and 5-8 lol
                writer=csv.writer(semestercsvfile)
                writer.writerow([f'{degreename}',"","","",f'{self.schoolname}',"The University of Texas at Austin"])
                writer.writerow(['','','','',''])
                # due to the complex nature of the rows this needs a csv object
                # this actually wont be necessary in excel. 
                writer.writerow(['','Course Code','Course Name','Hours','Category','Upper/Lower Division'])
                
                
                for semesternum in range(numberofsemesters):
                    semester=list(semesterdictionary)[semesternum]
                    # semester courses is a dictionary of its own as well
                    semestercourses=semesterdictionary[semester]
                    
                    csvobjectdict[rowcount]=[f'{semester}']
                    rowcount+=1
                    for coursenameindex in range(len(semestercourses)):

                        coursename=list(semestercourses)[coursenameindex]
                        # if its NOT a list of lists:
                        if len(semestercourses[coursename])==4 and not isinstance(semestercourses[coursename][0],list):
                            coursecode, coursehours, upperdivstatus, coursecategory=semestercourses[coursename]
                            csvobjectdict[rowcount]=["",coursecode,coursename.replace('removemelater',''),coursehours,coursecategory,upperdivstatus]
                            rowcount+=1
                            if coursehours!='' and "or" not in coursecode[0:2]:
                                totalhours+=int(coursehours)
                        
                        else:
                            listofcourses=semestercourses[coursename]
                            for i in range(len(listofcourses)):
                                try:
                                    coursecode, coursehours, upperdivstatus, coursecategory=listofcourses[i]
                                    csvobjectdict[rowcount]=["",coursecode,coursename.replace('removemelater',''),coursehours,coursecategory,upperdivstatus]
                                    rowcount+=1

                                    if coursehours!='' and "or" not in coursecode[0:2]:
                                        totalhours+=int(coursehours)
                                       
                                except ValueError:
                                    print(f'\n\nError for this courselist:\n {listofcourses}\n\n')



                    # line between semesters
                    csvobjectdict[rowcount]=['','','','','']
                    rowcount+=1




                print(f'len csv object={len(csvobjectdict)}')
                for row in csvobjectdict:
                    writer.writerow(csvobjectdict[row])

                writer.writerow(['','','',f'Total Hours: {totalhours}',''])
                writer.writerow(['DegreeView','','','','','DegreeView'])

                # this is for those giant architecture majors and some engineering that take like 6 years
        

                



        

        print(f'Made a semestercsv for {degreename}:\n As {semestercsvfilename}')
#-----------------------------------------------------------------------------------

     
    def make_greentheme_excel_files(
                self,
                schoolnamecolor="00CC66",
                bigbordercolor="66FF99",
                smallerbordercolor="99FFCC",
                subheadingbordercolor="66FF99",
                gridlinecolor="99FFCC",
                rowtextcolor="FFFFFF",
                titlecolor="FFFFFF",
                mainbackgroundcolor="004D26",
                paddingbackgroundcolor="007F3D",
                datafontname="Helvetica Neue",
                titlefontname="Playfair Display",
                logofontname="Raleway",
                logocolor="2F3E46"
    ):




        '''
        This uses the result of the scrape semester data function to make an excel file out of it.
        
        

        Keep for now

        '''
        for i in range(1,len(self.schooldata)):
            
            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Starting process for {degreename} excel file')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            # now use this to write to csv files.
            # I can actually finish this fast
            
            degreefolderpath=os.path.join(self.schoolfolderpath,degreename)
            # can change this quite easily
            # theres spaces which isnt great. But theres also spaces in the degreename. 


            degreenamecleaned=degreename.replace(' ','').lower().split('(')
            degreenamecleaned=degreenamecleaned[0]+"-"+degreenamecleaned[-1]
            degreenamecleaned=degreenamecleaned.replace(')','')

            # change the name of it here
            dark_semeseter_excel_filename=os.path.join(degreefolderpath,f'{degreenamecleaned}-darktheme-semesterfile.xlsx')

            totalhours=0
            numberofsemesters=len(semesterdictionary)
            excelobject=[]                        
            print(f'Excel object has been reset to {len(excelobject)}\n\n\n')



            '''
            Use Font and OpenPyXL to create nicely formatted tabular data
            '''

            semesterworkbook=Workbook()
            # gets the default worksheet
            ws=semesterworkbook.active



                
            titlefont=Font(size=26,bold=True,color=titlecolor,name=titlefontname)

            headingborder=Border(bottom=Side(style='mediumDashDot',color=smallerbordercolor))

            headingsfonts=Font(size=18,bold=True,color='ffffff')
            subheadingsfont=Font(size=17,bold=True,color='ffffff')
            # lighter burnt orange
            utnamefont=Font(size=19, color=schoolnamecolor, name='Georgia', bold=True)
            # update later
            semesterfont=Font(bold=True,size=18,color='ffffff')

            # applied to the actual semester data. 
            datafont=Font(size=15,name=datafontname,color=rowtextcolor)


            leftalign=Alignment(horizontal='left')
            centeralign=Alignment(horizontal='center')
            



            # total height of 40


            # first row after title that is. Start at row 3 since rows 1 and 2 were merged
            firstrow=[f'{degreename}',"","","",f'{self.schoolname}',"The University of Texas at Austin"]
                
            # colindex is like i            
            # start= 1 means start at the first column
            for col_index, value in enumerate(firstrow, start=1):
                # adjust row here
                headingcell=ws.cell(row=3, column=col_index, value=value)
                # use.font to assign the font I see
                if col_index==6:

                    headingcell.font=utnamefont
                    headingcell.border=headingborder
                elif col_index==1:
                    headingcell.font=headingsfonts
                    headingcell.border=headingborder
                    headingcell.alignment=centeralign
                else:
                    headingcell.font=headingsfonts
                    headingcell.border=headingborder

                    # print(datacell.font.size)


            blankrow=['','','','','','','','']
            # appends to the next empty row. Like writer.writerow(['']) for csvs
            ws.append(blankrow)
            
                        
            subheadingrow=['','Course Code','Course Name','Hours','Category','Upper/Lower Division']
            ws.append(subheadingrow)
            # now apply styles
            for cell in ws[5]:
                cell.font=subheadingsfont
                cell.alignment=leftalign


            # now the meat of the file, the data
            



            # for each semester
            rowcount=0
            for semesternum in range(numberofsemesters):
                semester=list(semesterdictionary)[semesternum]
                # semester courses is a dictionary of its own as well
                semestercourses=semesterdictionary[semester]
                # that means the semester will be on the first row
                excelobject.append([f'{semester}   '])
                rowcount+=1
                for coursenameindex in range(len(semestercourses)):

                    coursename=list(semestercourses)[coursenameindex]
                    # if its NOT a list of lists, ie, a normal course entry
                    if len(semestercourses[coursename])==4 and not isinstance(semestercourses[coursename][0],list):
                        coursecode, coursehours, upperdivstatus, coursecategory=semestercourses[coursename]
                        excelobject.append(["",coursecode,coursename,coursehours,coursecategory,upperdivstatus])
                        rowcount+=1
                        if coursehours!='':
                            totalhours+=int(coursehours)
                    
                    else:
                        listofcourses=semestercourses[coursename]
                        for i in range(len(listofcourses)):
                            coursecode, coursehours, upperdivstatus, coursecategory=listofcourses[i]
                            excelobject.append(["",coursecode,coursename,coursehours,coursecategory,upperdivstatus])
                            rowcount+=1

                            if coursehours!='':
                                totalhours+=int(coursehours)


                # line between semesters
                excelobject.append(['','','','',''])
                rowcount+=1

            # adding to excel file
            # rowval is important for adding rows in order
            rowval=6

            for rowentry in range(len(excelobject)):
                # update it here so it updates by row not column...although
                rowval += 1
                        # start at column one, and then remember the padding rows are added later. 
                for col_index, value in enumerate(excelobject[rowentry], start=1):
                    # change alignnment here. 1 indexed not 0
                    datacell = ws.cell(row=rowval, column=col_index, value=value)

                    if col_index==1:
                        # use.font to assign the font I see
                        datacell.font=semesterfont
                        datacell.alignment=Alignment(horizontal='right')

                    elif col_index in [2,3]:
                        datacell.font=datafont
                        datacell.alignment=leftalign
                    elif col_index==4: #hourscol
                        datacell.font=datafont
                        datacell.alignment=centeralign
                    elif col_index in [5,6]:
                        datacell.font=datafont
                        datacell.alignment=leftalign
                    
            
                       


            print(f'len excel object={len(excelobject)}, ie number of datarows')
            

# -----------------------------end data stuff --------------------------------------------------------
            ws.append(blankrow)
            totalhoursrow=['','','',f'Total Hours: {totalhours}','','']

            # this lastrow value is actually used for the last TWO rows
            lastrowindex=rowcount+6
            for col_index, value in enumerate(totalhoursrow, start=1):
                
                # use.font to assign the font I see
                cell=ws.cell(row=lastrowindex+1, column=col_index, value=value)
                # FF=full opacity 
                cell.font=datafont

           
            lastrow=['DegreeView','','','','','degreeviewsite.com']
            lastrowindex+=2 # 6 rows before we start data stuff. Then rowcount is the amount of data. 

            # change the logo colors here
            for col_index, value in enumerate(lastrow, start=1):
                
                # use.font to assign the font I see
                lastcell=ws.cell(row=lastrowindex, column=col_index, value=value)
                # FF=full opacity 
                if col_index==6:
                    # site link cell
                    lastcell.font=Font(name='Roboto',size=19, bold=True, color='e7e9eb')
                    lastcell.alignment=Alignment(horizontal='left',vertical='bottom')

                else:
                    # logo cell
                    lastcell.font=Font(name=logofontname,size=30, bold=True, color=logocolor)
                    lastcell.alignment=Alignment(horizontal='left',vertical='center')

                # one more cause now we wrote the actual last row there
                ws.row_dimensions[lastrowindex+1].height = 50



            # ---------------formatting stuff------------------
            
            # add padding
            
            
            ws.insert_rows(1)      #new row
            ws.insert_cols(1)       # new col

            # ws.columns gets columns. Each column is a tuple of cells
            for column_cells in ws.columns:
                # max function gets the max in a list. Same for min
                # generator must be in ()
                cellwidthgenerator=(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                colwidth = max(cellwidthgenerator)

                firstcell=column_cells[0]
                # every cell has a .column_letter attribute
                col_letter = firstcell.column_letter
                col_index=column_index_from_string(col_letter)
                # scale the width factor to make the columns wider
                if col_index==7:  
                    # make the UT Austin column alot wider
                    ws.column_dimensions[col_letter].width = int(colwidth)*2
                elif col_index==6:
                    ws.column_dimensions[col_letter].width = int(colwidth)*1.5
                elif col_index==5:
                    ws.column_dimensions[col_letter].width=int(colwidth)*.7
                elif col_index==3:
                    ws.column_dimensions[col_letter].width=int(colwidth)*2
                   
                # the semester column
                elif col_index==2:
                    ws.column_dimensions[col_letter].width=int(colwidth)*1.26
            

                else:
                    ws.column_dimensions[col_letter].width = int(colwidth)*1.2
                



            # set sizes for the padding row and column. Units of default font which is usually Arial it seems

            ws.column_dimensions['A'].width=15
            ws.column_dimensions['H'].width=15
            ws.row_dimensions[1].height=25
            ws.row_dimensions[lastrowindex+2].height=25

            # ------------------------------------TITLE STUFF-------------------------------------------

            # create a merged 'Degree view and degreename header can't lie".
            ws.merge_cells('B2:G3')    
            ws.row_dimensions[2].height = 30
            ws.row_dimensions[3].height = 30
            mergedrowcontent=f'{degreename} Sample Semester Layout'
            
            # refer to top left of merged cells
            titlecell=ws['B2']
            titlecell.value=mergedrowcontent
            titlecell.alignment = Alignment(horizontal='center', vertical='center')
            titlecell.font=titlefont

        # ------------------------ CONTINUE FORMATTING STUFF -----------------------------------------------------
            # apply a border around the entire file -----------------------------------------------------------------------------

            # make it responsive based on the amount of data
            # remove the +1 here to have no border
            rows = list(ws['B2':f'G{lastrowindex+1}'])
            min_row = 2
            max_row = lastrowindex+1

            min_col = column_index_from_string('B')  
            max_col = column_index_from_string('G') 

            
            # where border code used to be
            
            # make the entire worksheet a color:
            paddingbackgroundcolor=PatternFill(fill_type="solid", start_color=paddingbackgroundcolor) #end_color='0000FF' fill_type="gray125" or linear later

            # have the background be like a padding. 
            for row in ws.iter_rows(min_row=1, max_row=lastrowindex+2, min_col=1, max_col=8):
                for cell in row:
                    cell.fill = paddingbackgroundcolor
            
            # reverse the background for cells with content:

            backgroundcolor=PatternFill(fill_type="solid", start_color=mainbackgroundcolor)
            # re-add gridline borders:

            gridline_border = Border(
                # make gridline border not exactly white
                left=Side(border_style="thin", color=gridlinecolor),
                right=Side(border_style="thin", color=gridlinecolor),
                top=Side(border_style="thin", color=gridlinecolor),
                bottom=Side(border_style="thin", color=gridlinecolor)
            )

            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.fill = backgroundcolor
                    cell.border=gridline_border
            
            # this is the border that will go around everything.
            # Use logic to only add border to the cells on the outside.

            entire_ws_border=Side(style='thick',color=bigbordercolor)
            
            for row_index, row in enumerate(rows, start=min_row):
                for col_index, cell in enumerate(row, start=min_col):
                    current = cell.border

                    
                    # remember we defined side. and you do Border(Side=style))
                    '''
                    This is the final peice of the puzzle that adds borders.
                    If sides if empty, then
                    cell.border = Border(), which defines no borders. Normally it will be entire_ws_border border preset. 
                    '''
                    cell.border = Border(
            top=entire_ws_border if row_index == min_row else current.top,
            bottom=entire_ws_border if row_index == max_row else current.bottom,
            left=entire_ws_border if col_index == min_col else current.left,
            right=entire_ws_border if col_index == max_col else current.right
        )

            
            # where is this one?
            # this is like the subheading border
            for row in ws.iter_rows(min_row=6, max_row=6, min_col=3, max_col=7):
                for cell in row:
                    current = cell.border
                    cell.border = Border(
                        top=current.top,
                        bottom=Side(style='medium',color=subheadingbordercolor),  # Only change bottom keep thick border around everything
                        left=current.left,
                        right=current.right
                    )

            
            degreeviewpath='/Users/shalevwiden/Downloads/Projects/degreeview'
            semesterworkbook.save(dark_semeseter_excel_filename)
            print(f'Saved workbook at {dark_semeseter_excel_filename}')
    
     
    def automate_degreeplan_excel_files(self):




        '''
        This uses the result of the scrape semester data function to make an excel file out of it.
        For the CSV files I did them fully vertical. For these excel files, since I can control what columns and cells things are going,
        I want it to be more of a horizontal feel. 

        '''
        for i in range(1,len(self.schooldata)):
            
            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Starting process for {degreename} excel file')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            # now use this to write to csv files.
            # I can actually finish this fast
            
            degreefolderpath=os.path.join(self.schoolfolderpath,degreename)
            excelfolderpath=os.path.join(degreefolderpath,'excelfiles')
            if not os.path.exists(excelfolderpath):
                os.mkdir(excelfolderpath)

            # can change this quite easily
            # theres spaces which isnt great. But theres also spaces in the degreename. 


            degreenamecleaned=degreename.replace(' ','').lower().split('(')
            degreenamecleaned=degreenamecleaned[0]+"-"+degreenamecleaned[-1]
            degreenamecleaned=degreenamecleaned.replace(')','')

            # change the name of it here

            originalconfig=self.finishconfigpath('originalconfig.json')
            darkconfig=self.finishconfigpath('darkthemeconfig.json')
            blueconfig=self.finishconfigpath('colorconfigs/bluetheme.json')
            greenconfig=self.finishconfigpath('colorconfigs/green.json')
            desertconfig=self.finishconfigpath('themedconfigs/desert.json')
            greyscaleconfig=self.finishconfigpath('themedconfigs/greyscale.json')

            oceanconfig=self.finishconfigpath('themedconfigs/oceantheme.json')
            pastelconfig=self.finishconfigpath('themedconfigs/pasteltheme.json')
            primarycolorsconfig=self.finishconfigpath('themedconfigs/primarycolors.json')
            blackconfig=self.finishconfigpath('colorconfigs/black.json')

            neonconfig=self.finishconfigpath('themedconfigs/neon.json')



            def make_themed_file(configpath,theme):
                '''
                This is the most modular it can be.
                '''
                filename=f'{degreenamecleaned}-{theme.lower().replace(' ','').strip()}.xlsx'
                savepath=os.path.join(excelfolderpath,filename)
                with open(configpath,'r') as configjson:
                    configjson=json.load(configjson)
                
                config={
                    "savepath":savepath,
                    "universityname":self.universityname,
                    "schoolname":self.schoolname,
                    "degreename":degreename,
                    "semesterdictionary":semesterdictionary

                }
                config.update(configjson)

                self.make_degreeplan_excel_files(**config)

            make_themed_file(configpath=originalconfig,theme='originaltheme')

            make_themed_file(configpath=darkconfig,theme='darktheme')

            make_themed_file(configpath=greenconfig,theme='greentheme')
            
            make_themed_file(configpath=desertconfig,theme='desert')
            make_themed_file(configpath=blueconfig,theme='bluetheme')

            make_themed_file(configpath=greyscaleconfig,theme='greyscale')
            make_themed_file(configpath=oceanconfig,theme='ocean')
            make_themed_file(configpath=pastelconfig,theme='pastel')
            make_themed_file(configpath=primarycolorsconfig,theme='primarycolors')
            make_themed_file(configpath=blackconfig,theme='blacktheme')
            make_themed_file(configpath=neonconfig,theme='neon')





    def makehorizontalexcelfiles(self):
        '''Maybe do this later. With Excel it shouldnt be toooo hard'''
        
        pass
    
  



    def get_universitywide_stats(self):
        '''This makes the csv file. Then I'll still have to scrape data and make visualizations.'''

        degreenamelist=[]
        semestercountlist=[]
        totalhourslist=[]        
        numberofcourseslist=[]

        totalmajorhourslist=[]
        numberofmajorcourses=[]

        for i in range(1,len(self.schooldata)):

        # in this entire folder we are doing things by degree.


            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Getting stats for {degreename}')

            degreenamelist.append(degreename)
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            # now use this to write to csv files.
            # I can actually finish this fast
            
            universitystatsfolder="/Users/shalevwiden/Downloads/Projects/degreeview/stats"
            

            numberofsemesters=len(semesterdictionary)
            semestercountlist.append(numberofsemesters)

            totalhours=0

            coursecount=0
            majorcoursecount=0

            majorhourscount=0

            for semesternum in range(numberofsemesters):
                semester=list(semesterdictionary)[semesternum]
                # semester courses is a dictionary of its own as well
                semestercourses=semesterdictionary[semester]
                # that means the semester will be on the first row
                
                for coursenameindex in range(len(semestercourses)):

                    coursename=list(semestercourses)[coursenameindex]
                    # if its NOT a list of lists, ie, a normal course entry
                    if len(semestercourses[coursename])==4 and not isinstance(semestercourses[coursename][0],list):
                        coursecode, coursehours, upperdivstatus, coursecategory=semestercourses[coursename]
                        
                        coursecount+=1
                        if coursehours!='':
                            totalhours+=int(coursehours)
                            if 'major' in coursecategory.lower():
                                majorcoursecount+=1
                                majorhourscount+=int(coursehours)
                    
                    else:
                        listofcourses=semestercourses[coursename]
                        for i in range(len(listofcourses)):
                            coursecode, coursehours, upperdivstatus, coursecategory=listofcourses[i]
                            coursecount+=1

                            if coursehours!='':
                                totalhours+=int(coursehours)
                                if 'major' in coursecategory.lower():
                                    majorcoursecount+=1
                                    majorhourscount+=int(coursehours)

            # append all the data FOR ONE DEGREE to the lists which are FOR THE WHLE SCHOOl
            totalhourslist.append(totalhours)
            numberofcourseslist.append(coursecount)
            totalmajorhourslist.append(majorhourscount)
            numberofmajorcourses.append(majorcoursecount)




        utstatsfilepath=os.path.join(universitystatsfolder,'universityoftexas_stats.csv')

        with open(utstatsfilepath,'a') as utstatsfile:
            writer=csv.writer(utstatsfile)

            writer.writerow([f'{self.schoolname}'])

            # simply write all of the data
            # degreename, semestercount, totalhours, numberofcourses, majorhours, majorcourses
            for datarow in zip(degreenamelist,semestercountlist,totalhourslist,totalmajorhourslist,numberofcourseslist,numberofmajorcourses):
                writer.writerow(datarow)
            # cereates a blank line
            writer.writerow([])



            # can change this quite easily
            '''
            This is used to perform highest number of major hours and normal hours in datavisualizations.py
            '''
        return [degreenamelist,totalhourslist,totalmajorhourslist]


    def get_schoolspecific_stats():
        for i in range(1,len(self.schooldata)):

        # in this entire folder we are doing things by degree.


            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Getting stats for {degreename}')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            
            school_stats_folder=os.path.join(self.schoolfolderpath,f'{self.schoolname}stats')
            if not os.path.exists(school_stats_folder):
                os.mkdir(school_stats_folder) 


    def get_degree_per_school_counts(self):
        # returns a 2 item list
        return [self.schoolname,len(self.schooldata)-1]


    def get_degree_stats():
        '''
        Make txt files with this so I can easily read it and place it on the website. 
        '''
        pass
    def make_mpl_graphs():
        ''' Make some matplotlib graphs for each Degree as well.
            For organization
        '''
        pass













# archdata testing
def archtesting():
    archdata=theasset[0]

    archschoolkey=list(archdata)[0]
    archschoolname=archdata[archschoolkey]
    # print(f'School key: {archschoolkey}. School name: {archschoolname}')

    architecturefiles=makeSemesterFiles(schooldata=archdata)
    print('Testing:\nArchitecture School Folder Path')
    print(architecturefiles.schoolfolderpath)
    print('Testing:\nArchitecture School Name')

    print('Getting the attribute:\n')
    print(getattr(architecturefiles,'schoolname'))
    print()
    architecturefiles.upload_to_database()
# archtesting()


def make_universitywide_stats(theasset):

    universitystatsfolder="/Users/shalevwiden/Downloads/Projects/degreeview/stats"
    # if its deleted this will remake it

    if not os.path.exists( universitystatsfolder):
        os.mkdir( universitystatsfolder) 

    utstatsfilepath=os.path.join(universitystatsfolder,'universityoftexas_stats.csv')


    # this will rewrite it everytime I start the file. Clearing it


    with open(utstatsfilepath,'w') as utstatsfile:
        writer=csv.writer(utstatsfile)
        writer.writerow(['Degree Stats','','','','','The University of Texas at Austin'])
        writer.writerow(['','','','','',''])


        writer.writerow(['Degree Name','Number of Semesters','Total Hours','Total Major Hours','Number of Classes',"Number of Major Classes"])
        writer.writerow([])

    for schooldict in theasset:
        schoolobject=makeSemesterFiles(schooldata=schooldict)
        # this appends to the csv file
        schoolobject.get_universitywide_stats()
    
    # now add the closing row
    with open(utstatsfilepath,'a') as utstatsfile:
        writer=csv.writer(utstatsfile)
        writer.writerow(['','','','','',''])

        writer.writerow(['DegreeView','','','','',''])
    
    print('Hopefully made stats folder')

def make_big_datalist(theasset):
    '''
    This is used to find highest number of major hours and normal hours for a degree
    '''
    biglist=[]
    for schooldict in theasset:
        schoolobject=makeSemesterFiles(schooldata=schooldict)
        degreenamelist, totalhours,totalmajorhours=schoolobject.get_universitywide_stats()
        biglist.append([degreenamelist,totalhours,totalmajorhours])
    return biglist
# print(make_big_datalist(theasset=theasset))


def makecountlist(theasset):
    '''
    This counts the number of degree programs per school, allowing you to make a pie chart. 
    '''
    countlist=[]
    for schooldict in theasset:
        schoolobject=makeSemesterFiles(schooldata=schooldict)
        count=schoolobject.get_degree_per_school_counts()
        countlist.append(count)
    print(countlist)


def unpacktheasset_into_makefilesclass(theasset):
    '''
    This is the primary function that works with the functions in the class to make files
    '''
    
    for schooldict in theasset:
        schoolobject=makeSemesterFiles(schooldata=schooldict)
        schoolobject.upload_to_database()   
       
    


def main(theasset):
    for schooldict in theasset:
        schoolobject=makeSemesterFiles(schooldata=schooldict)
        

        # schoolobject.make_mermaid_files()
        # schoolobject.create_mmd_pdfs()
        # schoolo\bject.create_oddnumbered_mmd_pdfs()
        schoolobject.makecsvfiles()
        schoolobject.automate_degreeplan_excel_files()


main(theasset=theasset)
# for a later document where I do this same thing but replicated. For this I will simply just modify the "make_majorcourses_csvs.py"
