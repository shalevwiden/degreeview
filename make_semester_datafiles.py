import requests
import bs4
from bs4 import BeautifulSoup

import sys
import os
import subprocess
import random

import csv
import openpyxl
import json




from openpyxl import load_workbook
from openpyxl import Workbook
# use this to define the function define_start_column
from openpyxl.utils import column_index_from_string

# this assigns cells colors
from openpyxl.styles import PatternFill

import time
import os

from openpyxl.styles import Font
from openpyxl.styles import Border, Side, Alignment

# for working with pdfs:
import cairosvg
from cairosvg import svg2pdf

# png to pdf conversion
from PIL import Image


from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.colors import Color


# for merging the pdfs: and reading them to get data. 
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from pathlib import Path  


# for database stuff
import sqlite3


# Define the absolute path to the folder containing the file you want to import
# Replace this with your actual absolute path
pdftemplatesfolder = Path('/Users/shalevwiden/Downloads/Coding_Files/Python/BeautifulSoup_Library/PDF_Templates')

# Add the folder to the front of sys.path so Python will search here first
sys.path.insert(0, str(pdftemplatesfolder))

# Now you can import the function as if it were local
from stareoverlay import create_stare_png


exceltemplatespath = Path("/Users/shalevwiden/Downloads/Projects/dvwebsitecreation/sourcefiles/pythonfunctions")
sys.path.append(str(exceltemplatespath))

from excel_templates import make_checkerboardfile, make_excelfile




# good to check everythings working with the venv:
if __name__=='__main__':
    print(f'the version of beautiful soup is\n {(bs4.__version__)}')
    print(f'the version of requests is\n {(requests.__version__)}')
    print(f'the version of openpyxl is\n {(openpyxl.__version__)}')

    print(f'\nthe python version being used is:{sys.executable}\n')


# this script is gonna be.. huge. Yeah this is the behemoth one
# the asset is a LIST of dictionaries

with open('theassetcontainment.json') as assetjson:
    theasset=json.load(assetjson)

        
from scrapesemesterdata import getallcourses_splitbysemester


class makeSemesterFiles:
    '''
    Each object of this class should take an item (list item) from the asset.
    The asset of 0-14 for each school. 
    For testing we will use the asset of 0.
   
    '''
    # school data is like theasset[0], which is what each class object takes
    def __init__(self,schooldata):

        # you can print every attribute in the class which is the superpower tho. 

        # use self to make things an attribute
        self.schooldata=schooldata
        
        self.schoolnamekey=list(schooldata)[0]
        self.schoolname=schooldata[self.schoolnamekey]

        # this one is fixed 
        self.degreeviewfolderpath='/Users/shalevwiden/Downloads/Projects/originaldegreeview'
        # this should work. If not I need to find a mystery
        self.schoolfolderpath=os.path.join(self.degreeviewfolderpath,self.schoolname)
        
        self.schoolcolordict={'School of Architecture':'#cdbee9', 'Red McCombs School of Business':'#f9e4af', 
 'School of Civic Leadership':'#096c6c', 'Moody College of Communication':'#7e0107',
   'College of Education':'#86dff5', 'Cockrell School of Engineering':'#e35501', 
   'College of Fine Arts':"#633224", 'John A. and Katherine G. Jackson School of Geosciences':'#a3d29c', 
   'School of Information':'#2bc5e0', 
 'College of Liberal Arts':'#f3f3f3', 'College of Natural Sciences':'#f6c44b', 'School of Nursing':'#e3971f',
  'College of Pharmacy':"#BAB86C", 'Lyndon B. Johnson School of Public Affairs':'#096c6c', 'Steve Hicks School of Social Work':'#f0b420'}
        
        self.schoolcolor=self.schoolcolordict[self.schoolname]

        self.logopath='/Users/shalevwiden/Downloads/Projects/dvwebsitecreation/logo_design/logo6.png'


        # this might have to change later lol
        self.websiteurl_img_path='/Users/shalevwiden/Downloads/Projects/dvwebsitecreation/logo_design/websiteurl.png'

    def __str__(self):
        returnstring=f'This is an object to create files for {self.schoolname}'
        return returnstring
   
    def upload_to_database(self):
        '''
        
        Uses SQLite to upload to databases.
        We will then to quieries on said databases.

        
        '''
        for i in range(1,len(self.schooldata)):
            
            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Starting process for {degreename} ')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            # now use this to write to csv files.
            # I can actually finish this fast
            
            degreefolderpath=os.path.join(self.schoolfolderpath,degreename)
            # can change this quite easily


            # its a little different from the database name, use underscore instead of hyphen
            degreenamecleaned=degreename.replace(' ','').lower().split('(')
            degreenamecleaned=degreenamecleaned[0]+"-"+degreenamecleaned[-1]
            degreenamecleaned=degreenamecleaned.replace(')','')

            

            databasepath=os.path.join(degreefolderpath,f'{degreenamecleaned}-database.db')
            # hyphens and commas not allowed in tablename
            tabledegreename=degreenamecleaned.replace('-','_').replace(',','_').replace('&','and').replace("'","")

            tablename=f'{tabledegreename}_table'
            def maketable():
                '''This creates the table for course data in the db'''
                with sqlite3.connect(databasepath) as conn:
                    cursor=conn.cursor()

                # table name lowercase
                    createtablecommand=f'''CREATE TABLE IF NOT EXISTS "{tablename}" (Coursecode TEXT , Coursename TEXT, Hours INT, Category TEXT, UpperDivStatus TEXT );'''
                    cursor.execute(createtablecommand)
                    conn.commit()
            maketable()
            totalhours=0
            numberofsemesters=len(semesterdictionary)
            

                # .write stuff now mf
                # split up sems 1-4 and 5-8 lol
                
            with sqlite3.connect(databasepath) as conn:
                cursor=conn.cursor()

                    
                for semesternum in range(numberofsemesters):
                    semester=list(semesterdictionary)[semesternum]
                    # semester courses is a dictionary of its own as well
                    semestercourses=semesterdictionary[semester]
                        
                    for coursenameindex in range(len(semestercourses)):

                        coursename=list(semestercourses)[coursenameindex]
                            # if its NOT a list of lists:
                        if len(semestercourses[coursename])==4 and not isinstance(semestercourses[coursename][0],list):
                            coursecode, coursehours, upperdivstatus, coursecategory=semestercourses[coursename]
                            cursor.execute(f'INSERT INTO {tablename} (Coursecode, Coursename, Hours, Category, UpperDivStatus) values(?,?,?,?,?);',
                                   [coursecode,coursename,coursehours,coursecategory,upperdivstatus])
                            

                            if coursehours!='':
                                totalhours+=int(coursehours)
                            
                        else:
                            listofcourses=semestercourses[coursename]
                            for i in range(len(listofcourses)):
                                coursecode, coursehours, upperdivstatus, coursecategory=listofcourses[i]
                                cursor.execute(f'INSERT INTO {tablename} (Coursecode, Coursename, Hours, Category, UpperDivStatus) values(?,?,?,?,?);',
                                   [coursecode,coursename,coursehours,coursecategory,upperdivstatus])

                                if coursehours!='':
                                    totalhours+=int(coursehours)


                    # empty line between semesters
                    cursor.execute(f'INSERT INTO {tablename} (Coursecode, Coursename, Hours, Category, UpperDivStatus) values(?,?,?,?,?);',
                                   [None,None,None,None,None])
                                   
                 # commit to the database 
                conn.commit()
                print(f'Created {degreenamecleaned}-database.db\n')
                





            
            
    def makecsvfiles(self):
        '''
        This uses the result of the scrape semester data function to make a csv file out of it.
        '''

        # for each degree in school data
        for i in range(1,len(self.schooldata)):
            
            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Starting process for {degreename} ')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            # now use this to write to csv files.
            # I can actually finish this fast
            
            degreefolderpath=os.path.join(self.schoolfolderpath,degreename)
            # can change this quite easily


            degreenamecleaned=degreename.replace(' ','').lower().split('(')
            degreenamecleaned=degreenamecleaned[0]+"-"+degreenamecleaned[-1]
            degreenamecleaned=degreenamecleaned.replace(')','')

            semestercsvfilename=os.path.join(degreefolderpath,f'{degreenamecleaned}-semestercsvfile.csv')

            totalhours=0
            numberofsemesters=len(semesterdictionary)
            csvobjectdict={}                        
            print(f'\ncsv object has been rest to {len(csvobjectdict)}\n\n\n')
            
            rowcount=0
            with open(semestercsvfilename,'w',newline='') as semestercsvfile:

                # .write stuff now mf
                # split up sems 1-4 and 5-8 lol
                writer=csv.writer(semestercsvfile)
                writer.writerow([f'{degreename}',"","","",f'{self.schoolname}',"The University of Texas at Austin"])
                writer.writerow(['','','','',''])
                # due to the complex nature of the rows this needs a csv object
                # this actually wont be necessary in excel. 
                writer.writerow(['','Course Code','Course Name','Hours','Category','Upper/Lower Division'])
                
                
                for semesternum in range(numberofsemesters):
                    semester=list(semesterdictionary)[semesternum]
                    # semester courses is a dictionary of its own as well
                    semestercourses=semesterdictionary[semester]
                    
                    csvobjectdict[rowcount]=[f'{semester}']
                    rowcount+=1
                    for coursenameindex in range(len(semestercourses)):

                        coursename=list(semestercourses)[coursenameindex]
                        # if its NOT a list of lists:
                        if len(semestercourses[coursename])==4 and not isinstance(semestercourses[coursename][0],list):
                            coursecode, coursehours, upperdivstatus, coursecategory=semestercourses[coursename]
                            csvobjectdict[rowcount]=["",coursecode,coursename,coursehours,coursecategory,upperdivstatus]
                            rowcount+=1
                            if coursehours!='':
                                totalhours+=int(coursehours)
                        
                        else:
                            listofcourses=semestercourses[coursename]
                            for i in range(len(listofcourses)):
                                coursecode, coursehours, upperdivstatus, coursecategory=listofcourses[i]
                                csvobjectdict[rowcount]=["",coursecode,coursename,coursehours,coursecategory,upperdivstatus]
                                rowcount+=1

                                if coursehours!='':
                                    totalhours+=int(coursehours)


                    # line between semesters
                    csvobjectdict[rowcount]=['','','','','']
                    rowcount+=1




                print(f'len csv object={len(csvobjectdict)}')
                for row in csvobjectdict:
                    writer.writerow(csvobjectdict[row])

                writer.writerow(['','','',f'Total Hours: {totalhours}',''])
                writer.writerow(['DegreeView','','','','','DegreeView'])

                # this is for those giant architecture majors and some engineering that take like 6 years
        

                



        

        print(f'Made a semestercsv for {degreename}:\n As {semestercsvfilename}')
#-----------------------------------------------------------------------------------

        
    def make_excel_files(self):
        '''
        This uses the result of the scrape semester data function to make an excel file out of it.
        For the CSV files I did them fully vertical. For these excel files, since I can control what columns and cells things are going,
        I want it to be more of a horizontal feel. 

        '''
        for i in range(1,len(self.schooldata)):
            
            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Starting process for {degreename} excel file')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            # now use this to write to csv files.
            # I can actually finish this fast
            
            degreefolderpath=os.path.join(self.schoolfolderpath,degreename)
            # can change this quite easily
            # theres spaces which isnt great. But theres also spaces in the degreename. 


            degreenamecleaned=degreename.replace(' ','').lower().split('(')
            degreenamecleaned=degreenamecleaned[0]+"-"+degreenamecleaned[-1]
            degreenamecleaned=degreenamecleaned.replace(')','')

            semeseter_excel_filename=os.path.join(degreefolderpath,f'{degreenamecleaned}-semesterfile.xlsx')

            totalhours=0
            numberofsemesters=len(semesterdictionary)
            excelobject=[]                        
            print(f'Excel object has been reset to {len(excelobject)}\n\n\n')



            '''
            Use Font and OpenPyXL to create nicely formatted tabular data
            '''

            semesterworkbook=Workbook()
            # gets the default worksheet
            ws=semesterworkbook.active



                
            titlefont=Font(size=26,bold=True)

            headingborder=Border(bottom=Side(style='mediumDashDot'))

            headingsfonts=Font(size=18,bold=True)
            subheadingsfont=Font(size=17,bold=True)
            utnamefont=Font(size=19, color='FFBF5701', name='Georgia', bold=True)
            # update later
            semesterfont=Font(bold=True,size=18)

            # applied to the actual semester data. 
            datafont=Font(size=15,name='Helvetica')


            leftalign=Alignment(horizontal='left')
            centeralign=Alignment(horizontal='center')
            



            # total height of 40


            # first row after title that is. Start at row 3 since rows 1 and 2 were merged
            firstrow=[f'{degreename}',"","","",f'{self.schoolname}',"The University of Texas at Austin"]
                
            # colindex is like i            
            # start= 1 means start at the first column
            for col_index, value in enumerate(firstrow, start=1):
                # adjust row here
                headingcell=ws.cell(row=3, column=col_index, value=value)
                # use.font to assign the font I see
                if col_index==6:

                    headingcell.font=utnamefont
                    headingcell.border=headingborder
                elif col_index==1:
                    headingcell.font=headingsfonts
                    headingcell.border=headingborder
                    headingcell.alignment=centeralign
                else:
                    headingcell.font=headingsfonts
                    headingcell.border=headingborder

                    # print(datacell.font.size)


            blankrow=['','','','','','','','']
            # appends to the next empty row. Like writer.writerow(['']) for csvs
            ws.append(blankrow)
            
                        
            subheadingrow=['','Course Code','Course Name','Hours','Category','Upper/Lower Division']
            ws.append(subheadingrow)
            # now apply styles
            for cell in ws[5]:
                cell.font=subheadingsfont
                cell.alignment=leftalign


            # now the meat of the file, the data
            



            # for each semester
            rowcount=0
            for semesternum in range(numberofsemesters):
                semester=list(semesterdictionary)[semesternum]
                # semester courses is a dictionary of its own as well
                semestercourses=semesterdictionary[semester]
                # that means the semester will be on the first row
                excelobject.append([f'{semester}   '])
                rowcount+=1
                for coursenameindex in range(len(semestercourses)):

                    coursename=list(semestercourses)[coursenameindex]
                    # if its NOT a list of lists, ie, a normal course entry
                    if len(semestercourses[coursename])==4 and not isinstance(semestercourses[coursename][0],list):
                        coursecode, coursehours, upperdivstatus, coursecategory=semestercourses[coursename]
                        excelobject.append(["",coursecode,coursename,coursehours,coursecategory,upperdivstatus])
                        rowcount+=1
                        if coursehours!='':
                            totalhours+=int(coursehours)
                    
                    else:
                        listofcourses=semestercourses[coursename]
                        for i in range(len(listofcourses)):
                            coursecode, coursehours, upperdivstatus, coursecategory=listofcourses[i]
                            excelobject.append(["",coursecode,coursename,coursehours,coursecategory,upperdivstatus])
                            rowcount+=1

                            if coursehours!='':
                                totalhours+=int(coursehours)


                # line between semesters
                excelobject.append(['','','','',''])
                rowcount+=1

            # adding to excel file
            # rowval is important for adding rows in order
            rowval=6

            for rowentry in range(len(excelobject)):
                # update it here so it updates by row not column...although
                rowval += 1
                        # start at column one, and then remember the padding rows are added later. 
                for col_index, value in enumerate(excelobject[rowentry], start=1):
                    # change alignnment here. 1 indexed not 0
                    datacell = ws.cell(row=rowval, column=col_index, value=value)

                    if col_index==1:
                        # use.font to assign the font I see
                        datacell.font=semesterfont
                        datacell.alignment=Alignment(horizontal='right')

                    elif col_index in [2,3]:
                        datacell.font=datafont
                        datacell.alignment=leftalign
                    elif col_index==4: #hourscol
                        datacell.font=datafont
                        datacell.alignment=centeralign
                    elif col_index in [5,6]:
                        datacell.font=datafont
                        datacell.alignment=leftalign
                    
            
                       


            print(f'len excel object={len(excelobject)}, ie number of datarows')
            

# -----------------------------end data stuff --------------------------------------------------------
            ws.append(blankrow)
            totalhoursrow=['','','',f'Total Hours: {totalhours}','','']

            # this lastrow value is actually used for the last TWO rows
            lastrowindex=rowcount+6
            for col_index, value in enumerate(totalhoursrow, start=1):
                
                # use.font to assign the font I see
                cell=ws.cell(row=lastrowindex+1, column=col_index, value=value)
                # FF=full opacity 
                cell.font=datafont

           
            lastrow=['DegreeView','','','','','degreeviewsite.com']
            lastrowindex+=2 # 6 rows before we start data stuff. Then rowcount is the amount of data. 

            for col_index, value in enumerate(lastrow, start=1):
                
                # use.font to assign the font I see
                lastcell=ws.cell(row=lastrowindex, column=col_index, value=value)
                # FF=full opacity 
                if col_index==6:
                    lastcell.font=Font(name='Roboto',size=19, bold=True, color='000000')
                    lastcell.alignment=Alignment(horizontal='left',vertical='bottom')

                else:
                    # logo cell
                    lastcell.font=Font(name='Barlow',size=30, bold=True, color='0c48a5')
                    lastcell.alignment=Alignment(horizontal='left',vertical='center')

                # one more cause now we wrote the actual last row there
                ws.row_dimensions[lastrowindex+1].height = 50



            # ---------------formatting stuff------------------
            
            # add padding
            
            
            ws.insert_rows(1)      #new row
            ws.insert_cols(1)       # new col

            # ws.columns gets columns. Each column is a tuple of cells
            for column_cells in ws.columns:
                # max function gets the max in a list. Same for min
                # generator must be in ()
                cellwidthgenerator=(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                colwidth = max(cellwidthgenerator)

                firstcell=column_cells[0]
                # every cell has a .column_letter attribute
                col_letter = firstcell.column_letter
                col_index=column_index_from_string(col_letter)
                # scale the width factor to make the columns wider
                if col_index==7:  
                    # make the UT Austin column alot wider
                    ws.column_dimensions[col_letter].width = int(colwidth)*2
                elif col_index==6:
                    ws.column_dimensions[col_letter].width = int(colwidth)*1.5
                elif col_index==5:
                    ws.column_dimensions[col_letter].width=int(colwidth)*.7
                elif col_index==3:
                    ws.column_dimensions[col_letter].width=int(colwidth)*2
                   
                # the semester column
                elif col_index==2:
                    ws.column_dimensions[col_letter].width=int(colwidth)*1.26
            

                else:
                    ws.column_dimensions[col_letter].width = int(colwidth)*1.2
                



            # set sizes for the padding row and column. Units of default font which is usually Arial it seems

            ws.column_dimensions['A'].width=15
            ws.column_dimensions['H'].width=15
            ws.row_dimensions[1].height=25
            ws.row_dimensions[lastrowindex+2].height=25

            # ------------------------------------TITLE STUFF-------------------------------------------

            # create a merged 'Degree view and degreename header can't lie".
            ws.merge_cells('B2:G3')    
            ws.row_dimensions[2].height = 30
            ws.row_dimensions[3].height = 30
            mergedrowcontent=f'{degreename} Sample Semester Layout'
            
            # refer to top left of merged cells
            titlecell=ws['B2']
            titlecell.value=mergedrowcontent
            titlecell.alignment = Alignment(horizontal='center', vertical='center')
            titlecell.font=titlefont

        # ------------------------ CONTINUE FORMATTING STUFF -----------------------------------------------------
            # apply a border around the entire file -----------------------------------------------------------------------------

            # make it responsive based on the amount of data
            # remove the +1 here to have no border
            rows = list(ws['B2':f'G{lastrowindex+1}'])
            min_row = 2
            max_row = lastrowindex+1

            min_col = column_index_from_string('B')  
            max_col = column_index_from_string('G') 

            
            # where border code used to be
            
            # make the entire worksheet a color:
            backgroundcolor=PatternFill(fill_type="solid", start_color="F0F0F0") #end_color='0000FF' fill_type="gray125" or linear later

            # have the background be like a padding. 
            for row in ws.iter_rows(min_row=1, max_row=lastrowindex+2, min_col=1, max_col=8):
                for cell in row:
                    cell.fill = backgroundcolor
            
            # reverse the background for cells with content:

            whitefill=PatternFill(fill_type="solid", start_color="FFFFFF")
            # re-add gridline borders:

            gridline_border = Border(
                left=Side(border_style="thin", color="DDDDDD"),
                right=Side(border_style="thin", color="DDDDDD"),
                top=Side(border_style="thin", color="DDDDDD"),
                bottom=Side(border_style="thin", color="DDDDDD")
            )

            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.fill = whitefill
                    cell.border=gridline_border
            
            # this is the border that will go around everything.
            # Use logic to only add border to the cells on the outside.
            entire_ws_border=Side(style='thick',color='000000')
            
            for row_index, row in enumerate(rows, start=min_row):
                for col_index, cell in enumerate(row, start=min_col):
                    current = cell.border

                    
                    # remember we defined side. and you do Border(Side=style))
                    '''
                    This is the final peice of the puzzle that adds borders.
                    If sides if empty, then
                    cell.border = Border(), which defines no borders. Normally it will be entire_ws_border border preset. 
                    '''
                    cell.border = Border(
            top=entire_ws_border if row_index == min_row else current.top,
            bottom=entire_ws_border if row_index == max_row else current.bottom,
            left=entire_ws_border if col_index == min_col else current.left,
            right=entire_ws_border if col_index == max_col else current.right
        )

            
                    
            for row in ws.iter_rows(min_row=6, max_row=6, min_col=3, max_col=7):
                for cell in row:
                    current = cell.border
                    cell.border = Border(
                        top=current.top,
                        bottom=Side(style='medium'),  # Only change bottom keep thick border around everything
                        left=current.left,
                        right=current.right
                    )

            
            degreeviewpath='/Users/shalevwiden/Downloads/Projects/degreeview'
            semesterworkbook.save(semeseter_excel_filename)
            print(f'Saved workbook at {semeseter_excel_filename}')
    
    # dark theme now:


        
    def make_darktheme_excel_files(self):
        '''
        This uses the result of the scrape semester data function to make an excel file out of it.
        For the CSV files I did them fully vertical. For these excel files, since I can control what columns and cells things are going,
        I want it to be more of a horizontal feel. 

        '''
        for i in range(1,len(self.schooldata)):
            
            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Starting process for {degreename} excel file')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            # now use this to write to csv files.
            # I can actually finish this fast
            
            degreefolderpath=os.path.join(self.schoolfolderpath,degreename)
            # can change this quite easily
            # theres spaces which isnt great. But theres also spaces in the degreename. 


            degreenamecleaned=degreename.replace(' ','').lower().split('(')
            degreenamecleaned=degreenamecleaned[0]+"-"+degreenamecleaned[-1]
            degreenamecleaned=degreenamecleaned.replace(')','')

            # change the name of it here
            dark_semeseter_excel_filename=os.path.join(degreefolderpath,f'{degreenamecleaned}-darktheme-semesterfile.xlsx')

            totalhours=0
            numberofsemesters=len(semesterdictionary)
            excelobject=[]                        
            print(f'Excel object has been reset to {len(excelobject)}\n\n\n')



            '''
            Use Font and OpenPyXL to create nicely formatted tabular data
            '''

            semesterworkbook=Workbook()
            # gets the default worksheet
            ws=semesterworkbook.active



                
            titlefont=Font(size=26,bold=True,color='ffffff')

            headingborder=Border(bottom=Side(style='mediumDashDot',color='ffffff'))

            headingsfonts=Font(size=18,bold=True,color='ffffff')
            subheadingsfont=Font(size=17,bold=True,color='ffffff')
            # lighter burnt orange
            utnamefont=Font(size=19, color='ff9630', name='Georgia', bold=True)
            # update later
            semesterfont=Font(bold=True,size=18,color='ffffff')

            # applied to the actual semester data. 
            datafont=Font(size=15,name='Helvetica',color='ffffff')


            leftalign=Alignment(horizontal='left')
            centeralign=Alignment(horizontal='center')
            



            # total height of 40


            # first row after title that is. Start at row 3 since rows 1 and 2 were merged
            firstrow=[f'{degreename}',"","","",f'{self.schoolname}',"The University of Texas at Austin"]
                
            # colindex is like i            
            # start= 1 means start at the first column
            for col_index, value in enumerate(firstrow, start=1):
                # adjust row here
                headingcell=ws.cell(row=3, column=col_index, value=value)
                # use.font to assign the font I see
                if col_index==6:

                    headingcell.font=utnamefont
                    headingcell.border=headingborder
                elif col_index==1:
                    headingcell.font=headingsfonts
                    headingcell.border=headingborder
                    headingcell.alignment=centeralign
                else:
                    headingcell.font=headingsfonts
                    headingcell.border=headingborder

                    # print(datacell.font.size)


            blankrow=['','','','','','','','']
            # appends to the next empty row. Like writer.writerow(['']) for csvs
            ws.append(blankrow)
            
                        
            subheadingrow=['','Course Code','Course Name','Hours','Category','Upper/Lower Division']
            ws.append(subheadingrow)
            # now apply styles
            for cell in ws[5]:
                cell.font=subheadingsfont
                cell.alignment=leftalign


            # now the meat of the file, the data
            



            # for each semester
            rowcount=0
            for semesternum in range(numberofsemesters):
                semester=list(semesterdictionary)[semesternum]
                # semester courses is a dictionary of its own as well
                semestercourses=semesterdictionary[semester]
                # that means the semester will be on the first row
                excelobject.append([f'{semester}   '])
                rowcount+=1
                for coursenameindex in range(len(semestercourses)):

                    coursename=list(semestercourses)[coursenameindex]
                    # if its NOT a list of lists, ie, a normal course entry
                    if len(semestercourses[coursename])==4 and not isinstance(semestercourses[coursename][0],list):
                        coursecode, coursehours, upperdivstatus, coursecategory=semestercourses[coursename]
                        excelobject.append(["",coursecode,coursename,coursehours,coursecategory,upperdivstatus])
                        rowcount+=1
                        if coursehours!='':
                            totalhours+=int(coursehours)
                    
                    else:
                        listofcourses=semestercourses[coursename]
                        for i in range(len(listofcourses)):
                            coursecode, coursehours, upperdivstatus, coursecategory=listofcourses[i]
                            excelobject.append(["",coursecode,coursename,coursehours,coursecategory,upperdivstatus])
                            rowcount+=1

                            if coursehours!='':
                                totalhours+=int(coursehours)


                # line between semesters
                excelobject.append(['','','','',''])
                rowcount+=1

            # adding to excel file
            # rowval is important for adding rows in order
            rowval=6

            for rowentry in range(len(excelobject)):
                # update it here so it updates by row not column...although
                rowval += 1
                        # start at column one, and then remember the padding rows are added later. 
                for col_index, value in enumerate(excelobject[rowentry], start=1):
                    # change alignnment here. 1 indexed not 0
                    datacell = ws.cell(row=rowval, column=col_index, value=value)

                    if col_index==1:
                        # use.font to assign the font I see
                        datacell.font=semesterfont
                        datacell.alignment=Alignment(horizontal='right')

                    elif col_index in [2,3]:
                        datacell.font=datafont
                        datacell.alignment=leftalign
                    elif col_index==4: #hourscol
                        datacell.font=datafont
                        datacell.alignment=centeralign
                    elif col_index in [5,6]:
                        datacell.font=datafont
                        datacell.alignment=leftalign
                    
            
                       


            print(f'len excel object={len(excelobject)}, ie number of datarows')
            

# -----------------------------end data stuff --------------------------------------------------------
            ws.append(blankrow)
            totalhoursrow=['','','',f'Total Hours: {totalhours}','','']

            # this lastrow value is actually used for the last TWO rows
            lastrowindex=rowcount+6
            for col_index, value in enumerate(totalhoursrow, start=1):
                
                # use.font to assign the font I see
                cell=ws.cell(row=lastrowindex+1, column=col_index, value=value)
                # FF=full opacity 
                cell.font=datafont

           
            lastrow=['DegreeView','','','','','degreeviewsite.com']
            lastrowindex+=2 # 6 rows before we start data stuff. Then rowcount is the amount of data. 

            # change the logo colors here
            for col_index, value in enumerate(lastrow, start=1):
                
                # use.font to assign the font I see
                lastcell=ws.cell(row=lastrowindex, column=col_index, value=value)
                # FF=full opacity 
                if col_index==6:
                    # site link cell
                    lastcell.font=Font(name='Roboto',size=19, bold=True, color='e7e9eb')
                    lastcell.alignment=Alignment(horizontal='left',vertical='bottom')

                else:
                    # logo cell
                    lastcell.font=Font(name='Barlow',size=30, bold=True, color='a9d2f8')
                    lastcell.alignment=Alignment(horizontal='left',vertical='center')

                # one more cause now we wrote the actual last row there
                ws.row_dimensions[lastrowindex+1].height = 50



            # ---------------formatting stuff------------------
            
            # add padding
            
            
            ws.insert_rows(1)      #new row
            ws.insert_cols(1)       # new col

            # ws.columns gets columns. Each column is a tuple of cells
            for column_cells in ws.columns:
                # max function gets the max in a list. Same for min
                # generator must be in ()
                cellwidthgenerator=(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                colwidth = max(cellwidthgenerator)

                firstcell=column_cells[0]
                # every cell has a .column_letter attribute
                col_letter = firstcell.column_letter
                col_index=column_index_from_string(col_letter)
                # scale the width factor to make the columns wider
                if col_index==7:  
                    # make the UT Austin column alot wider
                    ws.column_dimensions[col_letter].width = int(colwidth)*2
                elif col_index==6:
                    ws.column_dimensions[col_letter].width = int(colwidth)*1.5
                elif col_index==5:
                    ws.column_dimensions[col_letter].width=int(colwidth)*.7
                elif col_index==3:
                    ws.column_dimensions[col_letter].width=int(colwidth)*2
                   
                # the semester column
                elif col_index==2:
                    ws.column_dimensions[col_letter].width=int(colwidth)*1.26
            

                else:
                    ws.column_dimensions[col_letter].width = int(colwidth)*1.2
                



            # set sizes for the padding row and column. Units of default font which is usually Arial it seems

            ws.column_dimensions['A'].width=15
            ws.column_dimensions['H'].width=15
            ws.row_dimensions[1].height=25
            ws.row_dimensions[lastrowindex+2].height=25

            # ------------------------------------TITLE STUFF-------------------------------------------

            # create a merged 'Degree view and degreename header can't lie".
            ws.merge_cells('B2:G3')    
            ws.row_dimensions[2].height = 30
            ws.row_dimensions[3].height = 30
            mergedrowcontent=f'{degreename} Sample Semester Layout'
            
            # refer to top left of merged cells
            titlecell=ws['B2']
            titlecell.value=mergedrowcontent
            titlecell.alignment = Alignment(horizontal='center', vertical='center')
            titlecell.font=titlefont

        # ------------------------ CONTINUE FORMATTING STUFF -----------------------------------------------------
            # apply a border around the entire file -----------------------------------------------------------------------------

            # make it responsive based on the amount of data
            # remove the +1 here to have no border
            rows = list(ws['B2':f'G{lastrowindex+1}'])
            min_row = 2
            max_row = lastrowindex+1

            min_col = column_index_from_string('B')  
            max_col = column_index_from_string('G') 

            
            # where border code used to be
            
            # make the entire worksheet a color:
            backgroundcolor=PatternFill(fill_type="solid", start_color="2e4053") #end_color='0000FF' fill_type="gray125" or linear later

            # have the background be like a padding. 
            for row in ws.iter_rows(min_row=1, max_row=lastrowindex+2, min_col=1, max_col=8):
                for cell in row:
                    cell.fill = backgroundcolor
            
            # reverse the background for cells with content:

            darkfill=PatternFill(fill_type="solid", start_color="273746")
            # re-add gridline borders:

            gridline_border = Border(
                # make gridline border not exactly white
                left=Side(border_style="thin", color="c9d8e7"),
                right=Side(border_style="thin", color="c9d8e7"),
                top=Side(border_style="thin", color="c9d8e7"),
                bottom=Side(border_style="thin", color="c9d8e7")
            )

            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.fill = darkfill
                    cell.border=gridline_border
            
            # this is the border that will go around everything.
            # Use logic to only add border to the cells on the outside.

            entire_ws_border=Side(style='thick',color='e2effd')
            
            for row_index, row in enumerate(rows, start=min_row):
                for col_index, cell in enumerate(row, start=min_col):
                    current = cell.border

                    
                    # remember we defined side. and you do Border(Side=style))
                    '''
                    This is the final peice of the puzzle that adds borders.
                    If sides if empty, then
                    cell.border = Border(), which defines no borders. Normally it will be entire_ws_border border preset. 
                    '''
                    cell.border = Border(
            top=entire_ws_border if row_index == min_row else current.top,
            bottom=entire_ws_border if row_index == max_row else current.bottom,
            left=entire_ws_border if col_index == min_col else current.left,
            right=entire_ws_border if col_index == max_col else current.right
        )

            
            # where is this one?
            # this is like the subheading border
            for row in ws.iter_rows(min_row=6, max_row=6, min_col=3, max_col=7):
                for cell in row:
                    current = cell.border
                    cell.border = Border(
                        top=current.top,
                        bottom=Side(style='medium',color='e2effd'),  # Only change bottom keep thick border around everything
                        left=current.left,
                        right=current.right
                    )

            
            degreeviewpath='/Users/shalevwiden/Downloads/Projects/degreeview'
            semesterworkbook.save(dark_semeseter_excel_filename)
            print(f'Saved workbook at {dark_semeseter_excel_filename}')
    
    def automate_degreeplan_excel_files(self):




        '''
        This uses the result of the scrape semester data function to make an excel file out of it.
        For the CSV files I did them fully vertical. For these excel files, since I can control what columns and cells things are going,
        I want it to be more of a horizontal feel. 

        '''
        for i in range(1,len(self.schooldata)):
            
            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Starting process for {degreename} excel file')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            # now use this to write to csv files.
            # I can actually finish this fast
            
            degreefolderpath=os.path.join(self.schoolfolderpath,degreename)
            # can change this quite easily
            # theres spaces which isnt great. But theres also spaces in the degreename. 


            degreenamecleaned=degreename.replace(' ','').lower().split('(')
            degreenamecleaned=degreenamecleaned[0]+"-"+degreenamecleaned[-1]
            degreenamecleaned=degreenamecleaned.replace(')','')

            # change the name of it here

            def make_themed_file(configpath,theme):
                '''
                This is the most modular it can be.
                '''
                filename=f'{degreenamecleaned}-{theme.lower().replace(' ','').strip()}.xlsx'
                savepath=os.path.join(degreefolderpath,filename)
                with open(configpath,'r') as configjson:
                    configjson=json.load(configjson)
                
                config={
                    "savepath":savepath,
                    "semesterdictionary":semesterdictionary,
                    "theme":theme,

                }
                config.update(configjson)

                self.make_degreeplan_excel_files(**config)

    



    def makehorizontalexcelfiles(self):
        '''Maybe do this later. With Excel it shouldnt be toooo hard'''
        
        pass
    
    def make_mermaid_files(self):

# we start at 1 because 0 is the school name 
        for i in range(1,len(self.schooldata)):

        # in this entire folder we are doing things by degree.


            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Starting process for {degreename} mermaid file')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            # now use this to write to csv files.
            # I can actually finish this fast
            
            degreefolderpath=os.path.join(self.schoolfolderpath,degreename)
            # can change this quite easily
            # theres spaces which isnt great. But theres also spaces in the degreename. 

            diagram_folder=os.path.join(degreefolderpath,'diagrams_and_mmdstuff')
            if not os.path.exists(diagram_folder):
                os.mkdir(diagram_folder)

            # this is reassigned later anyway
            mermaid_file_path=os.path.join(diagram_folder,f'{degreename} semesterdiagram.mmd')
            

            totalhours=0
            numberofsemesters=len(semesterdictionary)
            diagramcode = '''
        graph TD
            '''
            # top down graph
            # layout stuff? fixed, automatic. 

        # flowchart:
            # curve: basis

            hexpicker="#cf840c"
            

            # establish classes
            title_class_name='title_class'
            title_class='   \n  classDef title_class fill:#fff,stroke:#ffffff,stroke-width:5px,color:#ffffff,font-size:26px,font-weight:bold,width:260px,padding:20px,text-align:center,rx:20,ry:20'


            diagramcode+=title_class

            major_cat_class_name='major_cat_class'
            major_cat_class='   \nclassDef major_cat_class fill: #1d79db, stroke:#fff,stroke-width:4px,color:#FFF,font-weight:bold,rx:20,ry:20' 
            #stroke is border.
            diagramcode+=major_cat_class

            core_cat_class_name='core_cat_class'
            core_cat_class='    \nclassDef core_cat_class  fill:#BF5706, stroke:#fff,stroke-width:4px,color:#FFF,font-weight:bold;'
            diagramcode+=core_cat_class

            # I can manually make this bigger with spaces

            more_than_3_class='     \nclassDef more_than_3_class  fill:#f96, color:#FFF,font-weight:bold,stroke:#fff,stroke-width:4px;'


            elective_cat_class_name="elective_cat_class"
            elective_cat_class='    \nclassDef elective_cat_class  color:#FFF,font-weight:bold,fill:#579d42, stroke:#fff,stroke-width:4px;'
            diagramcode+=elective_cat_class

            gened_cat_class_name='gened_cat_class'
            gened_cat_class='    \nclassDef gened_cat_class color:#FFF,font-weight:bold,fill:#ae6455, stroke:#fff,stroke-width:4px;'
            diagramcode+=gened_cat_class


            opportunity_cat_class_name='opportunity_cat_class'
            opportunity_cat_class='    \nclassDef opportunity_cat_class color:#FFF,font-weight:bold,fill:#f2bb16, stroke:#fff,stroke-width:4px;'
            diagramcode+=opportunity_cat_class

            other_class_name='other_class'
            other_class= '  \nclassDef other_class color:#FFF,font-weight:bold,fill:#9cadb7, stroke:#fff,stroke-width:4px;'
            diagramcode+=other_class
            # 

        # at the end add everything connection. For now I define the subgraph. I'll be able to set arrow styles and such too.

            # use coursecount to make stuff invisible
            coursecount=0
            totalhours=0
            arrowcount=0
            # for each semester
            # do 1 and then +1 why?
            for semesternum in range(numberofsemesters):
                semester=list(semesterdictionary)[semesternum]
                semestercourses=semesterdictionary[semester]

                # the subgraph will have the text 1st semester, 2nd semester, etc but names semester1, semester2, etc
                diagramcode+=f'    \n\nsubgraph semester{semesternum+1}[{semester}]\n'

                
            
                
                # this is looping through the courses PER semester
                keeparrowlist=[]
                keeparrowlist=[item+2 for item in keeparrowlist]

                courseinsemestercount=0
                for coursenameindex in range(len(semestercourses)):

                    coursename=list(semestercourses)[coursenameindex]
                    # node name must be unique
                    nodename=f'course{courseinsemestercount}semester{semesternum}' #keep in mind its 0 indexed

                    # this is like course1semester1

                    # if its NOT a list of lists:
                    if len(semestercourses[coursename])==4 and not isinstance(semestercourses[coursename][0],list):
                        coursecode, coursehours, upperdivstatus, coursecategory=semestercourses[coursename]

                        # adding to mermaid logic--------------



                        # now set the brackets based on category. Square, round, rounded, triangle, etc. 
                        # change all the shape types later and add more to the classes as well.

                        # CSS is a walk in the park compared to this...god


                        # and give them classes. 
                        if 'major' in coursecategory.lower():
                            diagramcode+=f'\n{nodename}("{coursecode}")\n'

                            # define the class in the mmd file

                            diagramcode+=f'\nclass {nodename} {major_cat_class_name}\n'
                            

                        elif 'core' in coursecategory.lower() and "major" not in coursecategory.lower():
                            diagramcode+=f'{nodename}(["{coursecode}"])\n'

                            diagramcode+=f'class {nodename} {core_cat_class_name}\n'

                        elif 'general education' in coursecategory.lower():
                            diagramcode+=f'{nodename}["{coursecode}"]\n'
                            diagramcode+=f'class {nodename} {gened_cat_class_name}\n'



                        elif 'elective' in coursecategory.lower():
                            diagramcode+=f'{nodename}{{{{"{coursecode}"}}}}\n'

                            diagramcode+=f'class {nodename} {elective_cat_class_name}\n'


                        elif 'opportunity' in coursecategory.lower():
                            diagramcode+=f'{nodename}>"{coursecode}"]\n'
                            diagramcode+=f'class {nodename} {opportunity_cat_class_name}\n'

                        else:
                            diagramcode+=f'{nodename}[{coursecode}]\n'
                            diagramcode+=f'class {nodename} {other_class_name}\n'



                        # ----------------------------------------------------------
                        courseinsemestercount+=1
                        if coursehours!='':
                            totalhours+=int(coursehours)

                    
                    else:
                        listofcourses=semestercourses[coursename]
                        for i in range(len(listofcourses)):
                            coursecode, coursehours, upperdivstatus, coursecategory=listofcourses[i]
                            # this works cause even if list is 1 in length itll still add
                            nodename=f'course{courseinsemestercount}semester{semesternum}'
                            courseinsemestercount+=1
                            

                            # coursename is an empty string so use coursename here
                            if 'major' in coursecategory.lower():
                                diagramcode+=f'\n{nodename}("{coursename}")\n'

                            # define the class in the mmd file

                                diagramcode+=f'\nclass {nodename} {major_cat_class_name}\n'
                            

                            elif 'core' in coursecategory.lower() and "major" not in coursecategory.lower():
                                diagramcode+=f'{nodename}(["{coursename}"])\n'

                                diagramcode+=f'class {nodename} {core_cat_class_name}\n'

                            elif 'general education' in coursecategory.lower():
                                diagramcode+=f'{nodename}["{coursename}"]\n'
                                diagramcode+=f'class {nodename} {gened_cat_class_name}\n'



                            elif 'elective' in coursecategory.lower():
                                diagramcode+=f'{nodename}{{{{"{coursename}"}}}}\n'

                                diagramcode+=f'class {nodename} {elective_cat_class_name}\n'


                            elif 'opportunity' in coursecategory.lower():
                                diagramcode+=f'{nodename}>"{coursename}"]\n'
                                diagramcode+=f'class {nodename} {opportunity_cat_class_name}\n'

                            else:
                                diagramcode+=f'{nodename}[{coursename}]\n'
                                diagramcode+=f'class {nodename} {other_class_name}\n'




                            if coursehours!='':
                                totalhours+=int(coursehours)
                    coursecount+=courseinsemestercount
                # dealing with in subgraph stuff

                arrowmarker='---'
                if courseinsemestercount==4:
                    diagramcode+=f'course0semester{semesternum}{arrowmarker}course1semester{semesternum}\n'
                    diagramcode+=f'course2semester{semesternum}{arrowmarker}course3semester{semesternum}\n'
                    arrowcount+=2

                elif courseinsemestercount==5:
                    leftorright='left' if random.randint(0,1)==0 else 'right'
                    if leftorright=='left':
                        diagramcode+=f'course0semester{semesternum}{arrowmarker}course1semester{semesternum}{arrowmarker}course2semester{semesternum}\n'
                        diagramcode+=f'course3semester{semesternum}{arrowmarker}course4semester{semesternum}\n'
                        arrowcount+=3

                    else:
                        diagramcode+=f'course0semester{semesternum}{arrowmarker}course1semester{semesternum}\n'
                        diagramcode+=f'course2semester{semesternum}{arrowmarker}course3semester{semesternum}{arrowmarker}course4semester{semesternum}\n'
                        arrowcount+=3

                elif courseinsemestercount>=6:
                    diagramcode+=f'course0semester{semesternum}{arrowmarker}course1semester{semesternum}{arrowmarker}course2semester{semesternum}\n'
                    diagramcode+=f'course3semester{semesternum}{arrowmarker}course4semester{semesternum}{arrowmarker}course5semester{semesternum}\n'
                    arrowcount+=4

                    if courseinsemestercount==7:
                        diagramcode+=f'course6semester{semesternum}\n'


                    elif courseinsemestercount==8:
                        diagramcode+=f'course6semester{semesternum}\n'
                        diagramcode+=f'course7semester{semesternum}\n'


                    




                diagramcode+='\nend'
                diagramcode+='\n'
            # now connect the semesters:
            
            # see which semesters are on which sides
            firsthalf=numberofsemesters//2
            secondhalf=(numberofsemesters//2+firsthalf) if numberofsemesters%2==0 else ((numberofsemesters//2)+1)+firsthalf
            # change this 
            titlenode=f'\ntitlenode("{degreename} Semester Layout")\n'
            diagramcode+=titlenode
            diagramcode+=f'class titlenode {title_class_name}\n'

            # this connects to the right semester essentially
            diagramcode+=f'titlenode==>semester1\n'    
            # this should be respoonsive based on semester lengths

            diagramcode+=f'titlenode==>semester{secondhalf+1-firsthalf}\n'

            # make all arrows hidden except for the semester arrows...
            for arrowcount in range(arrowcount+2):
                diagramcode+=f'linkStyle {arrowcount} stroke:transparent,fill:transparent\n'

            
            # split up the halves        
            # 
            firsthalfconnected='\n  '
            keeparrowcount=0

            for semesternum in range(1,firsthalf+1):
                # changeline styles here 
                # for readability
                if semesternum==(firsthalf+1)-1:
                    firsthalfconnected+=f'semester{semesternum}=====>legend\n'
                else:
                    firsthalfconnected+=f'semester{semesternum}====>'
                    keeparrowcount+=1

            diagramcode+='%% firsthalfconnected:\n'
            diagramcode+=firsthalfconnected


            secondhalfconnected='\n '

            for semesternum in range(firsthalf+1,secondhalf+1):

                # changeline styles here 

                if semesternum==secondhalf+1-1:
                    secondhalfconnected+=f'semester{semesternum}=====>legend\n'
                else:
                    secondhalfconnected+=f'semester{semesternum}====>'
                    keeparrowcount+=1

            diagramcode+='%% secondhalfconnected:'

            diagramcode+=secondhalfconnected+'\n'

            # now hide the connecting to legend arrows:
            diagramcode+='%% Hide the connecting to legend arrows.\n'

            # this works...for some reason
            diagramcode+=f'linkStyle {arrowcount+firsthalf} stroke:transparent,fill:transparent\n'
            diagramcode+=f'linkStyle {arrowcount+keeparrowcount+2} stroke:transparent,fill:transparent\n'


            # legend and stats node
            # this one will be all flat on the bottom. 
            diagramcode+=f'    \n\nsubgraph legend["Course Legend"]\n'
            
            diagramcode+=f'''
        majornode("Major Category Courses")\n
        corenode(["Core Category Courses"])\n
        genednode["General Education Category Courses"]\n
        electivenode{{{{"Elective Category Courses"}}}}\n
        opportunitynode>"Opportunity Category Courses"]\n
        othernode["Other/Unspecified"]\n

        '''    
            diagramcode+='end\n'

            diagramcode+=f'\nclass majornode {major_cat_class_name}\n'
            diagramcode+=f'\nclass corenode {core_cat_class_name}\n'
            diagramcode+=f'\nclass genednode {gened_cat_class_name}\n'
            diagramcode+=f'\nclass electivenode {elective_cat_class_name}\n'
            diagramcode+=f'\n class opportunitynode {opportunity_cat_class_name}\n'
            diagramcode+=f'\n class othernode {other_class_name}'


            








        # --------------Now actually building the file and saving it -------------------------------------------------
            # replace this with the path to the schoolfolder in the right file


            degreenamecleaned=degreename.replace(' ','').lower().split('(')
            degreenamecleaned=degreenamecleaned[0]+"-"+degreenamecleaned[-1]
            degreenamecleaned=degreenamecleaned.replace(')','')

            mermaid_file_path=os.path.join(diagram_folder,f'{degreenamecleaned}-semesterdiagram.mmd')

            svg_savefilepath=os.path.join(diagram_folder,f'{degreenamecleaned}-semesterdiagram.svg')
            originaltheme_png_savefilepath=os.path.join(diagram_folder,f'{degreenamecleaned}-semesterdiagram.png')





            
            with open(mermaid_file_path, "w") as mermaidsemesterfile:
                mermaidsemesterfile.write(diagramcode)


            themepath=os.path.abspath('darktheme.json')

            lightthemeconfig="/Users/shalevwiden/Downloads/Coding_Files/Mermaid/Coursemmd/mermaid_config_files/light_congif_file.json"


            def rendermmd(createdpath,savepath,configpath):
                # adding a background here actually works

                subprocess.run(["mmdc", "-i", createdpath, "-o", savepath,"--configFile",configpath,"--scale=3"])
            

            # create both svg and png
            rendermmd(createdpath=mermaid_file_path,savepath=originaltheme_png_savefilepath,configpath=lightthemeconfig)
            rendermmd(createdpath=mermaid_file_path,savepath=svg_savefilepath,configpath=lightthemeconfig)

            print(f'Created {mermaid_file_path} and saved it at {originaltheme_png_savefilepath}')
            print(f'degreename is {degreename}')

            

    def create_mmd_pdfs(self):

        '''
        This takes the pdfs or svgs created by the above function, and turns them into stylized pdfs.

        Above is where background and node styling can take place.

        Here is where logo, images, sizing, and page appendation can take place.

        This one only creates EVEN numbered semester pdfs.
        '''
        for i in range(1,len(self.schooldata)):

        # in this entire folder we are doing things by degree.


            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()


            degreenamecleaned=degreename.replace(' ','').lower().split('(')
            degreenamecleaned=degreenamecleaned[0]+"-"+degreenamecleaned[-1]
            degreenamecleaned=degreenamecleaned.replace(')','')

            print(f'Starting process for {degreename} mermaid file')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            numberofsemesters=len(semesterdictionary)
            if numberofsemesters%2==0:


                
                # now use this to write to csv files.
                # I can actually finish this fast
                
                degreefolderpath=os.path.join(self.schoolfolderpath,degreename)
                # can change this quite easily
                # theres spaces which isnt great. But theres also spaces in the degreename. 

                diagram_folder=os.path.join(degreefolderpath,'diagrams_and_mmdstuff')
                
                originaltheme_png_file=os.path.join(diagram_folder,f'{degreenamecleaned}-semesterdiagram.png')

                legendonlypng_path='/Users/shalevwiden/Downloads/Coding_Files/Mermaid/Coursemmd/legendfolder/legendonly.png'

                originaltheme_pdf_file=os.path.join(diagram_folder,f'{degreenamecleaned}-semesterdiagram.pdf')


            # step one: convert pdfs to pngs.

                def png_to_pdf(png_path, pdf_path):
                    # Open the PNG and convert to RGB (PDFs can't handle alpha channels)
                    image = Image.open(png_path).convert("RGB")
                    image.save(pdf_path, "PDF")
                    print(f"Saved PDF to: {pdf_path}")
                
                # this is the diagram pdf.
                

                png_to_pdf(png_path=originaltheme_png_file,pdf_path=originaltheme_pdf_file)
                print(f'Created png to pdf for {originaltheme_pdf_file}')


                # This is the legend only pdf. The png is made in seperatelegend.py.

                def createmainoverlaypdf(diagrampath, outputpath):

                    # first get pdf data.
                    base_pdf = PdfReader(diagrampath)
                    page = base_pdf.pages[0]
                    width = float(page.mediabox.width)
                    height = float(page.mediabox.height)
                    print(f'width{width},height {height}')
                    # create pdf canvas
                    c = canvas.Canvas(outputpath,  pagesize=(width, height))  # (612 x 792 pt)


                    # if I ever change the logo I need to update this and the path.
                    logowidth=1622/3.85
                    logoheight=478/3.85

                    websiteurlwidth=2800/6.7
                    websiteurlheight=386/6.7

                    r = 12 / 255
                    g = 72 / 255
                    b = 165 / 255
                    # to actually use colors
                    # c.setFillColor(Color(r, g, b))
                    # Text color
                    c.setFillColor(Color(0, 0, 0))

                    c.setFont("Helvetica-Bold", 67)    

                    maxwidth=width-160
                    # Niiice now it is centered
                    headingtext1 = f"{degreename}"
                    headingtext2="Semester Layout"

                    text_width1 = c.stringWidth(headingtext1, "Helvetica-Bold", 67)
                    if text_width1>maxwidth:
                        font_size = 57
                        c.setFont("Helvetica-Bold", 57)

                        text_width1 = c.stringWidth(headingtext1, "Helvetica-Bold", font_size)
                        if text_width1>maxwidth:
                            c.setFont("Helvetica-Bold", 50)
                            font_size = 50


                            text_width1 = c.stringWidth(headingtext1, "Helvetica-Bold", font_size)



                    x_center1=(width - text_width1) / 2
                    c.drawString(x_center1, height-155, headingtext1)

                    c.setFont("Helvetica-Bold", 67)    

                    text_width2 = c.stringWidth(headingtext2, "Helvetica-Bold", 67)


                    x_center2=(width - text_width2) / 2
                    # No semester layout string I'm good

                    # c.drawString(x_center2, height-257, headingtext2)




                    # create white rectangle in main overlay. 
                    c.setFillColorRGB(1, 1, 1)  
                    c.rect(0, 0, width, 350, fill=1, stroke=0) 

                    

                    c.drawImage(self.logopath, 25, 9, width=logowidth, height=logoheight)
                    c.drawImage(self.websiteurl_img_path,width-(websiteurlwidth*1.06) , 26, width=websiteurlwidth, height=websiteurlheight)
                    # draw the link path

                    # draw the link as well in the bottom

                    c.save()
                    return width,height
                

                mainoverlaypath=os.path.join(diagram_folder,'mainoverlay.pdf')
                
                createmainoverlaypdf(diagrampath=originaltheme_pdf_file,outputpath=mainoverlaypath)


                print(f'Created mainoverlaypath at {mainoverlaypath}\n ')
                # now just merge them, along with the legend.
                mergedlegendpdf_path='/Users/shalevwiden/Downloads/Coding_Files/Mermaid/Coursemmd/legendfolder/mergedlegend.pdf'

                def merge_diagram_and_overlay_withlegend(pdfpath,mergedlegendpdfpath,mainoverlaypath,outputpath):
                    # well this finally works somewhat
                    diagram_pdf = PdfReader(pdfpath)
                    legendpdf=PdfReader(mergedlegendpdfpath)
                    mainoverlay_pdf = PdfReader(mainoverlaypath)
                    writer = PdfWriter()


                    diagram_page = diagram_pdf.pages[0]
                    mainoverlay_page = mainoverlay_pdf.pages[0]  
                    
                    diagram_page.merge_page(mainoverlay_page)   #  visual merge
                    writer.add_page(diagram_page)

                    legendpage=legendpdf.pages[0]


                    # Add merged page to writer. This line was causing problems but its good to know
                    writer.add_page(legendpage)


                    with open(outputpath, "wb") as finalpdffile:
                        writer.write(finalpdffile)

                        # output path is the final pdf file which will be downloaded

                finaloutputpath=os.path.join(diagram_folder,f'{degreenamecleaned}-semesterlayout.pdf')

                merge_diagram_and_overlay_withlegend(pdfpath=originaltheme_pdf_file,mergedlegendpdfpath=mergedlegendpdf_path, mainoverlaypath=mainoverlaypath,outputpath=finaloutputpath)

                print(f'Created final output path at {finaloutputpath}')



                def delete_unnessary_files():
                    os.remove(originaltheme_pdf_file)
                    os.remove(mainoverlaypath)
                    os.remove(originaltheme_png_file)
                delete_unnessary_files()

                # -----------------MAKE EMPTY NODES HERE
                activateemptynodes=True
                if activateemptynodes:

                    emptynodesfolder=os.path.join(diagram_folder,'emptynodesfolder')
                    if not os.path.exists(emptynodesfolder):
                        os.mkdir(emptynodesfolder) 

                    time.sleep(3)

                    def svg_to_pdf_emptynodes(svg_file, pdf_file):
                        svg2pdf(url=svg_file, write_to=pdf_file)

                    # this should save the empty nodes pdf in the respective folder. Then I'll 
                    # just have to run it through a stylizing function.
                    # This requires you to generate svg functions for every 
                    svg_savefilepath=os.path.join(diagram_folder,f'{degreenamecleaned}-semesterdiagram.svg')

                    empty_nodes_pdffile=os.path.join(emptynodesfolder,f'{degreenamecleaned}-emptynodes.pdf')

                    svg_to_pdf_emptynodes(svg_file=svg_savefilepath,pdf_file=empty_nodes_pdffile)
                    # now run pdf through all the formatting- define functions above.

                    def create_emptynodes_overlaypdf(diagrampath, outputpath):

                        # first get pdf data.
                        base_pdf = PdfReader(diagrampath)
                        page = base_pdf.pages[0]
                        width = float(page.mediabox.width)
                        height = float(page.mediabox.height)
                        print(f'width{width},height {height}')
                        # create pdf canvas
                        c = canvas.Canvas(outputpath,  pagesize=(width, height))  # (612 x 792 pt)


                        # if I ever change the logo I need to update this and the path.
                        logowidth=1622/6
                        logoheight=478/6

                        websiteurlwidth=2800/11
                        websiteurlheight=386/11

                        r = 12 / 255
                        g = 72 / 255
                        b = 165 / 255
                        # to actually use colors
                        # c.setFillColor(Color(r, g, b))
                        # Text color
                        c.setFillColor(Color(0, 0, 0))

                        c.setFont("Helvetica-Bold", 39)    

                        maxwidth=width-90
                        # Niiice now it is centered
                        headingtext1 = f"{degreename}"
                        headingtext2="Semester Layout"

                        text_width1 = c.stringWidth(headingtext1, "Helvetica-Bold", 39)
                        if text_width1>maxwidth:
                            font_size = 35
                            c.setFont("Helvetica-Bold", 35)

                            text_width1 = c.stringWidth(headingtext1, "Helvetica-Bold", font_size)
                            



                        x_center1=(width - text_width1) / 2
                        c.drawString(x_center1, height-75, headingtext1)

                        c.setFont("Helvetica-Bold", 39)    

                        text_width2 = c.stringWidth(headingtext2, "Helvetica-Bold", 39)


                        x_center2=(width - text_width2) / 2
                        # c.drawString(x_center2, height-134, headingtext2)




                        # create white rectangle in main overlay to cover the legend is the thing
                        c.setFillColorRGB(1, 1, 1)  
                        c.rect(0, 0, width, 250, fill=1, stroke=0) 

                        

                        c.drawImage(self.logopath, 25, 9, width=logowidth, height=logoheight)
                        c.drawImage(self.websiteurl_img_path,width-(websiteurlwidth*1.06) , 26, width=websiteurlwidth, height=websiteurlheight)
                        # draw the link path

                        # draw the link as well in the bottom

                        c.save()
                        return width,height
                    

                    emptynodesoverlaypath=os.path.join(emptynodesfolder,'emptynodesoverlay.pdf')
                    if "emptynodesmainoverlay.pdf" not in os.listdir(diagram_folder):
                        create_emptynodes_overlaypdf(diagrampath=empty_nodes_pdffile,outputpath=emptynodesoverlaypath)
                    else:
                        print('\nEmpty nodes overlay was already there so didnt create a new one\n')






                    mergedlegendpdf_path='/Users/shalevwiden/Downloads/Coding_Files/Mermaid/Coursemmd/legendfolder/mergedlegend.pdf'
                    finaloutput_emptynodes_path=os.path.join(emptynodesfolder,f'{degreenamecleaned}-semesterlayout_emptynodes.pdf')

                    merge_diagram_and_overlay_withlegend(pdfpath=empty_nodes_pdffile,mergedlegendpdfpath=mergedlegendpdf_path, mainoverlaypath=emptynodesoverlaypath,outputpath=finaloutput_emptynodes_path)

                    def delete_nodeunnessary_files():
                        # all these are used for the creation of the final pdf
                        os.remove(empty_nodes_pdffile)
                        os.remove(emptynodesoverlaypath)
                        os.remove(svg_savefilepath)

                    delete_nodeunnessary_files()

                    print(f'\nCreated empty nodes final output path at {finaloutput_emptynodes_path}.\n')


            

    def create_oddnumbered_mmd_pdfs(self):

        '''
        This takes the pdfs or svgs created by the above function, and turns them into stylized pdfs.

        Above is where background and node styling can take place.

        Here is where logo, images, sizing, and page appendation can take place.
          -------------------------------------------------
        This one only creates ODD numbered semester pdfs.

        Also to run this you must also run create_mmd_files because this function and the even one delete the pngs after theyre created to save storage.
        '''

        for i in range(1,len(self.schooldata)):

        # in this entire folder we are doing things by degree.


            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()


            degreenamecleaned=degreename.replace(' ','').lower().split('(')
            degreenamecleaned=degreenamecleaned[0]+"-"+degreenamecleaned[-1]
            degreenamecleaned=degreenamecleaned.replace(')','')

            print(f'Starting process for {degreename} mermaid file')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            numberofsemesters=len(semesterdictionary)

            if numberofsemesters%2!=0:

            
                # now use this to write to csv files.
                # I can actually finish this fast
                
                degreefolderpath=os.path.join(self.schoolfolderpath,degreename)
                # can change this quite easily
                # theres spaces which isnt great. But theres also spaces in the degreename. 

                diagram_folder=os.path.join(degreefolderpath,'diagrams_and_mmdstuff')
                
                originaltheme_png_file=os.path.join(diagram_folder,f'{degreenamecleaned}-semesterdiagram.png')

                legendonlypng_path='/Users/shalevwiden/Downloads/Coding_Files/Mermaid/Coursemmd/legendfolder/legendonly.png'

                originaltheme_pdf_file=os.path.join(diagram_folder,f'{degreenamecleaned}-semesterdiagram.pdf')


            # step one: convert pdfs to pngs.

                def png_to_pdf(png_path, pdf_path):
                    # Open the PNG and convert to RGB (PDFs can't handle alpha channels)
                    image = Image.open(png_path).convert("RGB")
                    image.save(pdf_path, "PDF")
                    print(f"Saved PDF to: {pdf_path}")
                
                # this is the diagram pdf.
                

                png_to_pdf(png_path=originaltheme_png_file,pdf_path=originaltheme_pdf_file)
                print(f'Created png to pdf for {originaltheme_pdf_file}')


                # This is the legend only pdf. The png is made in seperatelegend.py.

                def createmainoverlaypdf(diagrampath, outputpath):

                    # first get pdf data.
                    base_pdf = PdfReader(diagrampath)
                    page = base_pdf.pages[0]
                    width = float(page.mediabox.width)
                    height = float(page.mediabox.height)
                    print(f'width{width},height {height}')
                    # create pdf canvas
                    c = canvas.Canvas(outputpath,  pagesize=(width, height))  # (612 x 792 pt)


                    # if I ever change the logo I need to update this and the path.
                    logowidth=1622/3.85
                    logoheight=478/3.85

                    websiteurlwidth=2800/6.7
                    websiteurlheight=386/6.7

                    r = 12 / 255
                    g = 72 / 255
                    b = 165 / 255
                    # to actually use colors
                    # c.setFillColor(Color(r, g, b))
                    # Text color
                    c.setFillColor(Color(0, 0, 0))

                    c.setFont("Helvetica-Bold", 67)    

                    maxwidth=width*.6
                    
                    headingtext1 = f"{degreename.split('(')[0]}"
                    headingtext2 = f"({degreename.split('(')[1]}"

                    headingtext3="Semester Layout"

                    # text_width1 = c.stringWidth(headingtext1, "Helvetica-Bold", 67)
                    # if text_width1>maxwidth:
                    #     font_size = 57
                    #     c.setFont("Helvetica-Bold", 57)

                    #     text_width1 = c.stringWidth(headingtext1, "Helvetica-Bold", font_size)
                    #     if text_width1>maxwidth:
                    #         c.setFont("Helvetica-Bold", 50)
                    #         font_size = 50


                    #         text_width1 = c.stringWidth(headingtext1, "Helvetica-Bold", font_size)



                    
                    x_position1=(190)
                    c.drawString(x_position1, height-455, headingtext1)


                    text_width2 = c.stringWidth(headingtext2, "Helvetica-Bold", 67)


                    x_position2=(190)
                    c.drawString(x_position2, height-550, headingtext2)

                    # heading text 2 will always be 67, "Semester Layout"
                    c.setFont("Helvetica-Bold", 67)

                    x_position3=(190)
                    # c.drawString(x_position3, height-645, headingtext3)




                    # create white rectangle in main overlay. 
                    c.setFillColorRGB(1, 1, 1)  
                    c.rect(0, 0, width, 350, fill=1, stroke=0) 

                    

                    c.drawImage(self.logopath, 25, 9, width=logowidth, height=logoheight)
                    c.drawImage(self.websiteurl_img_path,width-(websiteurlwidth*1.06) , 26, width=websiteurlwidth, height=websiteurlheight)
                    # draw the link path

                    # draw the link as well in the bottom

                    c.save()
                    return width,height
                

                mainoverlaypath=os.path.join(diagram_folder,'mainoverlay.pdf')
                
                createmainoverlaypdf(diagrampath=originaltheme_pdf_file,outputpath=mainoverlaypath)


                print(f'Created mainoverlaypath at {mainoverlaypath}\n ')
                # now just merge them, along with the legend.
                mergedlegendpdf_path='/Users/shalevwiden/Downloads/Coding_Files/Mermaid/Coursemmd/legendfolder/mergedlegend.pdf'

                def merge_diagram_and_overlay_withlegend(pdfpath,mergedlegendpdfpath,mainoverlaypath,outputpath):
                    # well this finally works somewhat
                    diagram_pdf = PdfReader(pdfpath)
                    legendpdf=PdfReader(mergedlegendpdfpath)
                    mainoverlay_pdf = PdfReader(mainoverlaypath)
                    writer = PdfWriter()


                    diagram_page = diagram_pdf.pages[0]
                    mainoverlay_page = mainoverlay_pdf.pages[0]  
                    
                    diagram_page.merge_page(mainoverlay_page)   #  visual merge
                    writer.add_page(diagram_page)

                    legendpage=legendpdf.pages[0]


                    # Add merged page to writer. This line was causing problems but its good to know
                    writer.add_page(legendpage)


                    with open(outputpath, "wb") as finalpdffile:
                        writer.write(finalpdffile)

                        # output path is the final pdf file which will be downloaded

                finaloutputpath=os.path.join(diagram_folder,f'{degreenamecleaned}-semesterlayout.pdf')

                merge_diagram_and_overlay_withlegend(pdfpath=originaltheme_pdf_file,mergedlegendpdfpath=mergedlegendpdf_path, mainoverlaypath=mainoverlaypath,outputpath=finaloutputpath)

                print(f'Created final output path at {finaloutputpath}')



                def delete_unnessary_files():
                    os.remove(originaltheme_pdf_file)
                    os.remove(mainoverlaypath)
                    os.remove(originaltheme_png_file)
                delete_unnessary_files()

                # -----------------MAKE EMPTY NODES HERE
                activateemptynodes=True
                if activateemptynodes:

                    emptynodesfolder=os.path.join(diagram_folder,'emptynodesfolder')
                    if not os.path.exists(emptynodesfolder):
                        os.mkdir(emptynodesfolder) 

                    time.sleep(3)

                    def svg_to_pdf_emptynodes(svg_file, pdf_file):
                        svg2pdf(url=svg_file, write_to=pdf_file)

                    # this should save the empty nodes pdf in the respective folder. Then I'll 
                    # just have to run it through a stylizing function.
                    # This requires you to generate svg functions for every 
                    svg_savefilepath=os.path.join(diagram_folder,f'{degreenamecleaned}-semesterdiagram.svg')

                    empty_nodes_pdffile=os.path.join(emptynodesfolder,f'{degreenamecleaned}-emptynodes.pdf')

                    svg_to_pdf_emptynodes(svg_file=svg_savefilepath,pdf_file=empty_nodes_pdffile)
                    # now run pdf through all the formatting- define functions above.

                    def create_emptynodes_overlaypdf(diagrampath, outputpath):

                        # first get pdf data.
                        base_pdf = PdfReader(diagrampath)
                        page = base_pdf.pages[0]
                        width = float(page.mediabox.width)
                        height = float(page.mediabox.height)
                        print(f'width{width},height {height}')
                        # create pdf canvas
                        c = canvas.Canvas(outputpath,  pagesize=(width, height))  # (612 x 792 pt)


                        # if I ever change the logo I need to update this and the path.
                        logowidth=1622/6
                        logoheight=478/6

                        websiteurlwidth=2800/11
                        websiteurlheight=386/11

                        r = 12 / 255
                        g = 72 / 255
                        b = 165 / 255
                        # to actually use colors
                        # c.setFillColor(Color(r, g, b))
                        # Text color
                        c.setFillColor(Color(0, 0, 0))

                        c.setFont("Helvetica-Bold", 39)    

                        maxwidth=width*.45
                        # Niiice now it is centered
                            
                        headingtext1 = f"{degreename.split('(')[0]}"
                        headingtext2 = f"({degreename.split('(')[1]}"

                        headingtext3="Semester Layout"

                        # smaller font since the svg is alot smaller
                        text_width1 = c.stringWidth(headingtext1, "Helvetica-Bold", 39)
                        if text_width1>maxwidth:
                            font_size = 35
                            c.setFont("Helvetica-Bold", 35)

                            text_width1 = c.stringWidth(headingtext1, "Helvetica-Bold", font_size)
                            


# --------------------------------------------

                        x_position1=(150)
                        c.drawString(x_position1, height-230, headingtext1)

                        c.setFont("Helvetica-Bold", 39)    

                        text_width2 = c.stringWidth(headingtext2, "Helvetica-Bold", 39)


                        x_position2=(150)
                        c.drawString(x_position2, height-290, headingtext2)



                        x_position3=(150)
                        # c.drawString(x_position3, height-350, headingtext3)


                        # create white rectangle in main overlay to cover the legend is the thing
                        c.setFillColorRGB(1, 1, 1)  
                        c.rect(0, 0, width, 250, fill=1, stroke=0) 

                        

                        c.drawImage(self.logopath, 25, 9, width=logowidth, height=logoheight)
                        c.drawImage(self.websiteurl_img_path,width-(websiteurlwidth*1.06) , 26, width=websiteurlwidth, height=websiteurlheight)
                        # draw the link path

                        # draw the link as well in the bottom

                        c.save()
                        return width,height
                    

                    emptynodesoverlaypath=os.path.join(emptynodesfolder,'emptynodesoverlay.pdf')
                    if "emptynodesmainoverlay.pdf" not in os.listdir(diagram_folder):
                        create_emptynodes_overlaypdf(diagrampath=empty_nodes_pdffile,outputpath=emptynodesoverlaypath)
                    else:
                        print('\nEmpty nodes overlay was already there so didnt create a new one\n')






                    mergedlegendpdf_path='/Users/shalevwiden/Downloads/Coding_Files/Mermaid/Coursemmd/legendfolder/mergedlegend.pdf'
                    finaloutput_emptynodes_path=os.path.join(emptynodesfolder,f'{degreenamecleaned}-semesterlayout_emptynodes.pdf')

                    merge_diagram_and_overlay_withlegend(pdfpath=empty_nodes_pdffile,mergedlegendpdfpath=mergedlegendpdf_path, mainoverlaypath=emptynodesoverlaypath,outputpath=finaloutput_emptynodes_path)

                    def delete_nodeunnessary_files():
                        # all these are used for the creation of the final pdf
                        os.remove(empty_nodes_pdffile)
                        os.remove(emptynodesoverlaypath)
                        os.remove(svg_savefilepath)

                    delete_nodeunnessary_files()

                    print(f'\nCreated empty nodes final output path at {finaloutput_emptynodes_path}.\n')







    def get_universitywide_stats(self):
        '''This makes the csv file. Then I'll still have to scrape data and make visualizations.'''

        degreenamelist=[]
        semestercountlist=[]
        totalhourslist=[]        
        numberofcourseslist=[]

        totalmajorhourslist=[]
        numberofmajorcourses=[]

        for i in range(1,len(self.schooldata)):

        # in this entire folder we are doing things by degree.


            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Getting stats for {degreename}')

            degreenamelist.append(degreename)
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            # now use this to write to csv files.
            # I can actually finish this fast
            
            universitystatsfolder="/Users/shalevwiden/Downloads/Projects/degreeview/stats"
            

            numberofsemesters=len(semesterdictionary)
            semestercountlist.append(numberofsemesters)

            totalhours=0

            coursecount=0
            majorcoursecount=0

            majorhourscount=0

            for semesternum in range(numberofsemesters):
                semester=list(semesterdictionary)[semesternum]
                # semester courses is a dictionary of its own as well
                semestercourses=semesterdictionary[semester]
                # that means the semester will be on the first row
                
                for coursenameindex in range(len(semestercourses)):

                    coursename=list(semestercourses)[coursenameindex]
                    # if its NOT a list of lists, ie, a normal course entry
                    if len(semestercourses[coursename])==4 and not isinstance(semestercourses[coursename][0],list):
                        coursecode, coursehours, upperdivstatus, coursecategory=semestercourses[coursename]
                        
                        coursecount+=1
                        if coursehours!='':
                            totalhours+=int(coursehours)
                            if 'major' in coursecategory.lower():
                                majorcoursecount+=1
                                majorhourscount+=int(coursehours)
                    
                    else:
                        listofcourses=semestercourses[coursename]
                        for i in range(len(listofcourses)):
                            coursecode, coursehours, upperdivstatus, coursecategory=listofcourses[i]
                            coursecount+=1

                            if coursehours!='':
                                totalhours+=int(coursehours)
                                if 'major' in coursecategory.lower():
                                    majorcoursecount+=1
                                    majorhourscount+=int(coursehours)

            # append all the data FOR ONE DEGREE to the lists which are FOR THE WHLE SCHOOl
            totalhourslist.append(totalhours)
            numberofcourseslist.append(coursecount)
            totalmajorhourslist.append(majorhourscount)
            numberofmajorcourses.append(majorcoursecount)




        utstatsfilepath=os.path.join(universitystatsfolder,'universityoftexas_stats.csv')

        with open(utstatsfilepath,'a') as utstatsfile:
            writer=csv.writer(utstatsfile)

            writer.writerow([f'{self.schoolname}'])

            # simply write all of the data
            # degreename, semestercount, totalhours, numberofcourses, majorhours, majorcourses
            for datarow in zip(degreenamelist,semestercountlist,totalhourslist,totalmajorhourslist,numberofcourseslist,numberofmajorcourses):
                writer.writerow(datarow)
            # cereates a blank line
            writer.writerow([])



            # can change this quite easily
            '''
            This is used to perform highest number of major hours and normal hours in datavisualizations.py
            '''
        return [degreenamelist,totalhourslist,totalmajorhourslist]


    def get_schoolspecific_stats():
        for i in range(1,len(self.schooldata)):

        # in this entire folder we are doing things by degree.


            key=list(self.schooldata)[i]
            degreename=key
            degreename=degreename.replace('/','-').strip()

            print(f'Getting stats for {degreename}')
            
            # kept as sugg link as continuity from make_makorcourses_csvs
            sugglink=self.schooldata[key]

            
            semesterdictionary=getallcourses_splitbysemester(suggcourse_link=sugglink)
            
            school_stats_folder=os.path.join(self.schoolfolderpath,f'{self.schoolname}stats')
            if not os.path.exists(school_stats_folder):
                os.mkdir(school_stats_folder) 


    def get_degree_per_school_counts(self):
        # returns a 2 item list
        return [self.schoolname,len(self.schooldata)-1]


    def get_degree_stats():
        '''
        Make txt files with this so I can easily read it and place it on the website. 
        '''
        pass
    def make_mpl_graphs():
        ''' Make some matplotlib graphs for each Degree as well.
            For organization
        '''
        pass













# archdata testing
def archtesting():
    archdata=theasset[0]

    archschoolkey=list(archdata)[0]
    archschoolname=archdata[archschoolkey]
    # print(f'School key: {archschoolkey}. School name: {archschoolname}')

    architecturefiles=makeSemesterFiles(schooldata=archdata)
    print('Testing:\nArchitecture School Folder Path')
    print(architecturefiles.schoolfolderpath)
    print('Testing:\nArchitecture School Name')

    print('Getting the attribute:\n')
    print(getattr(architecturefiles,'schoolname'))
    print()
    architecturefiles.upload_to_database()
# archtesting()


def make_universitywide_stats(theasset):

    universitystatsfolder="/Users/shalevwiden/Downloads/Projects/degreeview/stats"
    # if its deleted this will remake it

    if not os.path.exists( universitystatsfolder):
        os.mkdir( universitystatsfolder) 

    utstatsfilepath=os.path.join(universitystatsfolder,'universityoftexas_stats.csv')


    # this will rewrite it everytime I start the file. Clearing it


    with open(utstatsfilepath,'w') as utstatsfile:
        writer=csv.writer(utstatsfile)
        writer.writerow(['Degree Stats','','','','','The University of Texas at Austin'])
        writer.writerow(['','','','','',''])


        writer.writerow(['Degree Name','Number of Semesters','Total Hours','Total Major Hours','Number of Classes',"Number of Major Classes"])
        writer.writerow([])

    for schooldict in theasset:
        schoolobject=makeSemesterFiles(schooldata=schooldict)
        # this appends to the csv file
        schoolobject.get_universitywide_stats()
    
    # now add the closing row
    with open(utstatsfilepath,'a') as utstatsfile:
        writer=csv.writer(utstatsfile)
        writer.writerow(['','','','','',''])

        writer.writerow(['DegreeView','','','','',''])
    
    print('Hopefully made stats folder')

def make_big_datalist(theasset):
    '''
    This is used to find highest number of major hours and normal hours for a degree
    '''
    biglist=[]
    for schooldict in theasset:
        schoolobject=makeSemesterFiles(schooldata=schooldict)
        degreenamelist, totalhours,totalmajorhours=schoolobject.get_universitywide_stats()
        biglist.append([degreenamelist,totalhours,totalmajorhours])
    return biglist
# print(make_big_datalist(theasset=theasset))


def makecountlist(theasset):
    '''
    This counts the number of degree programs per school, allowing you to make a pie chart. 
    '''
    countlist=[]
    for schooldict in theasset:
        schoolobject=makeSemesterFiles(schooldata=schooldict)
        count=schoolobject.get_degree_per_school_counts()
        countlist.append(count)
    print(countlist)


def unpacktheasset_into_makefilesclass(theasset):
    '''
    This is the primary function that works with the functions in the class to make files
    '''
    
    for schooldict in theasset:
        schoolobject=makeSemesterFiles(schooldata=schooldict)
        # schoolobject.upload_to_database()   
       
    
# unpacktheasset_into_makefilesclass(theasset=theasset)


def runentirefile(theasset):

    for schooldict in theasset:
        schoolobject=makeSemesterFiles(schooldata=schooldict)
        

        schoolobject.make_greentheme_excel_files()

runentirefile(theasset=theasset)
# for a later document where I do this same thing but replicated. For this I will simply just modify the "make_majorcourses_csvs.py"
class makeMajorFiles:
    def __init__(self):
        pass